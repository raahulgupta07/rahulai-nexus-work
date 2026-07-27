"""
excel_structizer.py — Standalone structural parser for messy .xlsx workbooks.

Turns real-world spreadsheets (merged headers, title banners, totals rows,
stacked mini-tables, UI-metadata columns) into clean, queryable table dicts.

Layers implemented here (deterministic, openpyxl + pandas only):

  L0  Read the FULL cell grid via openpyxl (data_only=True) and capture
      ``ws.merged_cells.ranges``.  pandas is deliberately NOT used for structure
      because ``read_excel`` silently discards merges and guesses a single header.

  L1  Deterministic cleanup:
        1. unmerge   — propagate each merged range's top-left value into every
                       covered cell (so spanned headers/labels become real data).
        2. split     — cut a sheet into table blocks on fully-blank rows.
        3. title/junk— drop spanned "title" banner rows.
        4. header    — data starts at the first numeric-bearing row; every
                       (non-title) row above it is the header band.  Multi-row
                       header bands are merged column-wise by joining with "_".
        5. prune     — drop Total/Grand Total/Subtotal rows, fully-blank columns,
                       Unnamed/blank-header columns, and any ``*_SD_*`` UI column.
        6. clean     — snake_case, deduped, unique column names; trim to the data
                       rectangle.
        7. score     — confidence in [0,1].  High = one clear header + rectangular
                       data.  Low (< 0.6) = merged-heavy / headerless / ambiguous,
                       which signals the optional L2 (LLM) fallback.

  L2  Optional LLM fallback (only when L1 confidence < 0.6).  If a sibling
      ``llm_structizer`` module (same CONTRACT) is present it is used via its
      offline ``mock_llm`` path.  The import is guarded so this module runs
      fully standalone when that sibling does not exist yet.

Public API:  ``parse_workbook(path) -> list[dict]``  (see CONTRACT.md).
"""

from __future__ import annotations

import re
import warnings

import openpyxl
from openpyxl.utils import get_column_letter

# openpyxl warns loudly about unsupported extensions (data validation, etc.);
# they are irrelevant to structure extraction, so silence them.
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Tunables
# ---------------------------------------------------------------------------
CONFIDENCE_L2_THRESHOLD = 0.6      # below this we consult the L2 fallback
MIN_DATA_ROWS = 1                  # a block needs at least this many data rows
MIN_COLS = 1                       # ...and at least this many kept columns
DROP_BLOCK_BELOW_CONF = 0.35       # blocks weaker than this are treated as
                                   # knowledge (skipped) rather than junk tables
MAX_HEADER_ROWS = 3                # multi-row headers deeper than this are not
                                   # real headers (prose/records mistaken as such)
GLOSSARY_MAX_CONF = 0.30           # cap for non-tabular prose/glossary blocks

_TOTALS_RE = re.compile(r"^\s*(grand\s+total|sub\s*total|subtotal|total)\s*:?\s*$", re.I)
_SD_META_RE = re.compile(r"_sd_", re.I)          # UI metadata columns to drop
_UNNAMED_RE = re.compile(r"^unnamed[:\s_]*\d*$", re.I)


# ---------------------------------------------------------------------------
# Cell primitives
# ---------------------------------------------------------------------------
def _is_blank(v) -> bool:
    """True for empty cells (None or whitespace-only strings)."""
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _is_number(v) -> bool:
    """True for real numeric cells (int/float, but not bool, not NaN)."""
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        # guard against NaN
        return v == v
    return False


def _is_text(v) -> bool:
    return isinstance(v, str) and v.strip() != "" and not _is_number(v)


def _cell_text(v) -> str:
    """Flatten a cell to a trimmed single-line string for header building."""
    if v is None:
        return ""
    if isinstance(v, str):
        return re.sub(r"\s+", " ", v).strip()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ---------------------------------------------------------------------------
# L0 — grid + merges
# ---------------------------------------------------------------------------
def _load_grid(ws):
    """Return (grid, merged_ranges) for a worksheet.

    grid is a dense list[list] of raw values (None for empty), sized to
    ws.max_row x ws.max_column.  merged_ranges is the list of openpyxl ranges.
    """
    n_rows = ws.max_row or 0
    n_cols = ws.max_column or 0
    grid = [[None] * n_cols for _ in range(n_rows)]
    for row in ws.iter_rows():
        for cell in row:
            r, c = cell.row - 1, cell.column - 1
            if 0 <= r < n_rows and 0 <= c < n_cols:
                grid[r][c] = cell.value
    merged = list(ws.merged_cells.ranges)
    return grid, merged


def _unmerge(grid, merged):
    """Propagate every merged range's top-left value into all covered cells.

    Mutates a *copy* and returns it, so the raw grid is preserved.
    """
    out = [list(row) for row in grid]
    n_rows = len(out)
    n_cols = len(out[0]) if out else 0
    for rng in merged:
        r0, c0 = rng.min_row - 1, rng.min_col - 1
        r1, c1 = rng.max_row - 1, rng.max_col - 1
        if not (0 <= r0 < n_rows and 0 <= c0 < n_cols):
            continue
        val = out[r0][c0]
        for r in range(r0, min(r1 + 1, n_rows)):
            for c in range(c0, min(c1 + 1, n_cols)):
                out[r][c] = val
    return out


# ---------------------------------------------------------------------------
# Block detection
# ---------------------------------------------------------------------------
def _row_blank(grid, r, c0, c1):
    return all(_is_blank(grid[r][c]) for c in range(c0, c1 + 1))


def _col_blank(grid, c, r0, r1):
    return all(_is_blank(grid[r][c]) for r in range(r0, r1 + 1))


def _sheet_extent(grid):
    """Bounding box of all non-blank cells: (r0, r1, c0, c1) or None if empty."""
    rows_hit, cols_hit = [], []
    for r, row in enumerate(grid):
        for c, v in enumerate(row):
            if not _is_blank(v):
                rows_hit.append(r)
                cols_hit.append(c)
    if not rows_hit:
        return None
    return min(rows_hit), max(rows_hit), min(cols_hit), max(cols_hit)


def _split_row_blocks(grid, r0, r1, c0, c1):
    """Split the row range [r0,r1] into blocks separated by fully-blank rows.

    Returns list of (rstart, rend) inclusive row ranges (non-blank content).
    """
    blocks = []
    cur = None
    for r in range(r0, r1 + 1):
        if _row_blank(grid, r, c0, c1):
            if cur is not None:
                blocks.append(cur)
                cur = None
        else:
            if cur is None:
                cur = [r, r]
            else:
                cur[1] = r
    if cur is not None:
        blocks.append(cur)
    return [tuple(b) for b in blocks]


def _block_col_extent(grid, r0, r1, c0, c1):
    """Tighten a block's column window to columns that hold any content."""
    cols = [c for c in range(c0, c1 + 1) if not _col_blank(grid, c, r0, r1)]
    if not cols:
        return c0, c1
    return min(cols), max(cols)


# ---------------------------------------------------------------------------
# Title-row detection
# ---------------------------------------------------------------------------
def _is_title_row(grid, r, c0, c1):
    """A banner/title row: its non-blank cells are all the SAME value (an
    unmerged span) covering most of the width, or a single wide text cell.
    """
    vals = [grid[r][c] for c in range(c0, c1 + 1)]
    nonblank = [v for v in vals if not _is_blank(v)]
    width = c1 - c0 + 1
    if not nonblank:
        return False
    # single lone text cell across a wide row -> title
    if len(nonblank) == 1 and width >= 3 and _is_text(nonblank[0]):
        return True
    # spanned banner: every non-blank cell identical, covering >= half the width,
    # and textual (numbers spanning a row is not a title)
    first = _cell_text(nonblank[0])
    if (
        len(nonblank) >= max(2, width // 2)
        and all(_cell_text(v) == first for v in nonblank)
        and not all(_is_number(v) for v in nonblank)
    ):
        return True
    return False


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------
def _is_header_ish_row(grid, r, c0, c1):
    """A row that looks like column labels rather than a data record.

    Header labels are: textual, carry NO real numbers, are reasonably filled,
    and are short (not paragraph-length prose).  This is what lets us bound a
    multi-row header band and stop at the first data/prose row.
    """
    vals = [grid[r][c] for c in range(c0, c1 + 1)]
    nonblank = [v for v in vals if not _is_blank(v)]
    if len(nonblank) < 1:
        return False
    if any(_is_number(v) for v in nonblank):
        return False                      # numbers => data record, not header
    text = [v for v in nonblank if _is_text(v)]
    if len(text) < len(nonblank):         # dates / non-text => not a label row
        return False
    avg_len = sum(len(_cell_text(v)) for v in text) / len(text)
    if avg_len > 40:                      # paragraph prose => data, not header
        return False
    return True


def _detect_header(grid, r0, r1, c0, c1):
    """Return (title_rows, header_rows, data_r0, data_r1).

    - title_rows : leading banner rows to drop
    - header_rows: the (possibly multi-row) header band, capped at MAX_HEADER_ROWS
    - data range : [data_r0, data_r1] inclusive
    header_rows == [] means a headerless block.
    """
    # 1. peel leading title/banner rows
    r = r0
    title_rows = []
    while r <= r1 and _is_title_row(grid, r, c0, c1):
        title_rows.append(r)
        r += 1
    body_r0 = r
    if body_r0 > r1:
        return title_rows, [], body_r0, r1

    # 2. header band = leading run of "header-ish" label rows, bounded so that
    #    prose/records are never absorbed into the header.
    band = []
    r = body_r0
    while (
        r <= r1
        and len(band) < MAX_HEADER_ROWS
        and _is_header_ish_row(grid, r, c0, c1)
    ):
        band.append(r)
        r += 1
    data_r0 = r

    # need at least one data row under the header; otherwise headerless
    if not band or data_r0 > r1:
        return title_rows, [], body_r0, r1

    return title_rows, band, data_r0, r1


# ---------------------------------------------------------------------------
# Column name building
# ---------------------------------------------------------------------------
def _snake(name: str) -> str:
    name = _cell_text(name)
    name = name.lower()
    name = re.sub(r"[^0-9a-z]+", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    return name


def _dedupe(names):
    """Ensure unique names, appending _2, _3 ... to collisions."""
    seen = {}
    out = []
    for n in names:
        if n not in seen:
            seen[n] = 1
            out.append(n)
        else:
            seen[n] += 1
            out.append(f"{n}_{seen[n]}")
    return out


def _build_column(grid, header_rows, c):
    """Join a column's header cells (top -> bottom) into one raw label,
    collapsing consecutive duplicates produced by unmerged spans."""
    parts = []
    for r in header_rows:
        t = _cell_text(grid[r][c])
        if t and (not parts or parts[-1].lower() != t.lower()):
            parts.append(t)
    return " ".join(parts)


# ---------------------------------------------------------------------------
# Block extraction (the L1 core)
# ---------------------------------------------------------------------------
def _extract_block(grid, r0, r1, c0, c1, sheet_name, merged, used_names):
    """Turn one content block into a table dict (or None if not tabular)."""
    notes = []

    c0, c1 = _block_col_extent(grid, r0, r1, c0, c1)
    title_rows, header_rows, d0, d1 = _detect_header(grid, r0, r1, c0, c1)

    if title_rows:
        notes.append(f"dropped {len(title_rows)} title/banner row(s)")

    headerless = not header_rows
    if headerless:
        notes.append("no header detected (headerless block)")

    # --- assemble candidate columns -----------------------------------------
    raw_cols = []
    for c in range(c0, c1 + 1):
        raw = _build_column(grid, header_rows, c) if header_rows else ""
        raw_cols.append((c, raw))

    if len(header_rows) > 1:
        notes.append(f"merged {len(header_rows)}-row header")

    # --- data rows ----------------------------------------------------------
    data_rows_idx = list(range(d0, d1 + 1))
    # drop fully-blank data rows
    data_rows_idx = [r for r in data_rows_idx if not _row_blank(grid, r, c0, c1)]

    # drop Total / Subtotal / Grand Total rows (by first non-blank text cell)
    kept_rows = []
    dropped_totals = 0
    for r in data_rows_idx:
        first_txt = ""
        for c in range(c0, c1 + 1):
            if not _is_blank(grid[r][c]):
                first_txt = _cell_text(grid[r][c])
                break
        if _TOTALS_RE.match(first_txt):
            dropped_totals += 1
            continue
        kept_rows.append(r)
    if dropped_totals:
        notes.append(f"dropped {dropped_totals} totals row(s)")

    if len(kept_rows) < MIN_DATA_ROWS:
        return None  # nothing tabular here

    # --- decide which columns to keep ---------------------------------------
    keep_cols = []          # list of (col_idx, final_name, was_named)
    dropped_blank = dropped_meta = dropped_unnamed = 0
    for idx, (c, raw) in enumerate(raw_cols):
        col_blank = all(_is_blank(grid[r][c]) for r in kept_rows)
        # UI-metadata columns (…_SD_…) always go
        if raw and _SD_META_RE.search(raw):
            dropped_meta += 1
            continue
        if raw and _UNNAMED_RE.match(raw.strip()):
            dropped_unnamed += 1
            continue
        if col_blank and not raw:
            dropped_blank += 1
            continue
        if col_blank and raw:
            # header exists but no data below -> also junk
            dropped_blank += 1
            continue
        was_named = bool(raw)
        name = _snake(raw) if raw else ""
        if not name:
            name = f"column_{idx + 1}"
        keep_cols.append((c, name, was_named))

    if dropped_blank:
        notes.append(f"dropped {dropped_blank} blank column(s)")
    if dropped_meta:
        notes.append(f"dropped {dropped_meta} _SD_ metadata column(s)")
    if dropped_unnamed:
        notes.append(f"dropped {dropped_unnamed} unnamed column(s)")

    if len(keep_cols) < MIN_COLS:
        return None

    final_names = _dedupe([n for (_c, n, _w) in keep_cols])

    # --- build rows ---------------------------------------------------------
    rows = []
    for r in kept_rows:
        rows.append([grid[r][c] for (c, _n, _w) in keep_cols])

    # --- drop non-tabular fragments -----------------------------------------
    # A lone 1x1 / single-value scrap (e.g. a merged "Comment" box) is not a
    # table.  Require some 2-D substance.
    if len(keep_cols) == 1 and len(rows) < 2:
        return None

    # --- confidence ---------------------------------------------------------
    kept_col_idx = [c for (c, _n, _w) in keep_cols]

    # distinctness of the (named) columns: repeated names come from a spanned
    # banner mis-read as a header -> strong signal the block is not tabular.
    base_names = [re.sub(r"_\d+$", "", n) for (_c, n, _w) in keep_cols]
    distinct_ratio = len(set(base_names)) / len(base_names) if base_names else 0.0
    named_ratio = (
        sum(1 for (_c, _n, w) in keep_cols if w) / len(keep_cols)
        if keep_cols else 0.0
    )

    total_cells = len(rows) * len(keep_cols)
    filled = sum(1 for row in rows for v in row if not _is_blank(v))
    data_fill = filled / total_cells if total_cells else 0.0

    # merges are expected in real multi-row headers, so keep this penalty light
    merge_overlap = _count_merges_in(merged, r0, r1, c0, c1)
    merge_penalty = min(0.20, merge_overlap / 40.0)

    single_header_bonus = 1.0 if len(header_rows) == 1 else 0.0

    if headerless:
        conf = 0.20 + 0.25 * data_fill
    else:
        # reward distinct, well-named columns over a filled data rectangle
        conf = (
            0.40 * (0.5 * named_ratio + 0.5 * distinct_ratio)
            + 0.35 * data_fill
            + 0.15 * single_header_bonus
            + 0.10
        )
    conf = max(0.0, min(1.0, conf - merge_penalty))

    # --- non-tabular prose / glossary detection -----------------------------
    # No numbers anywhere + narrow + long-text cells == a term/definition or
    # free-text block, i.e. knowledge, not a queryable table.
    has_number = any(
        _is_number(grid[r][c])
        for r in kept_rows for c in kept_col_idx
    )
    text_cells = [
        _cell_text(grid[r][c])
        for r in kept_rows for c in kept_col_idx
        if not _is_blank(grid[r][c])
    ]
    mean_len = (sum(len(t) for t in text_cells) / len(text_cells)) if text_cells else 0
    if (not has_number) and len(keep_cols) <= 2 and mean_len > 25:
        conf = min(conf, GLOSSARY_MAX_CONF)
        notes.append("likely glossary/free-text (non-tabular)")

    # --- range string -------------------------------------------------------
    cell_range = "%s%d:%s%d" % (
        get_column_letter(min(kept_col_idx) + 1), r0 + 1,
        get_column_letter(max(kept_col_idx) + 1), d1 + 1,
    )

    name = _unique_table_name(sheet_name, used_names)

    return {
        "name": name,
        "sheet": sheet_name,
        "cell_range": cell_range,
        "columns": final_names,
        "rows": rows,
        "confidence": round(conf, 3),
        "method": "L1",
        "notes": notes,
    }


def _count_merges_in(merged, r0, r1, c0, c1):
    n = 0
    for rng in merged:
        if rng.min_row - 1 <= r1 and rng.max_row - 1 >= r0 and \
           rng.min_col - 1 <= c1 and rng.max_col - 1 >= c0:
            n += 1
    return n


# ---------------------------------------------------------------------------
# Naming helpers
# ---------------------------------------------------------------------------
def _unique_table_name(sheet_name, used_names):
    base = _snake(sheet_name) or "table"
    name = base
    i = 2
    while name in used_names:
        name = f"{base}_{i}"
        i += 1
    used_names.add(name)
    return name


# ---------------------------------------------------------------------------
# L2 fallback (optional, guarded)
# ---------------------------------------------------------------------------
def _maybe_l2(grid, r0, r1, c0, c1, l1_table):
    """If the sibling llm_structizer exists, try its offline mock path for a
    low-confidence block.  Never raises: returns l1_table on any problem."""
    try:
        # Resolve the sibling both as a top-level module (standalone use) and as
        # a package module (inside the app: app.services.llm_structizer).
        try:
            import llm_structizer  # standalone / same-dir import
        except Exception:
            from app.services import llm_structizer  # packaged import
    except Exception:
        return l1_table  # standalone mode — no L2 available

    try:
        sub = [row[c0:c1 + 1] for row in grid[r0:r1 + 1]]
        grid_sample = llm_structizer.build_grid_sample(sub)
        spec = llm_structizer.mock_llm(grid_sample)
        tables = llm_structizer.apply_spec(sub, spec)
        if not tables:
            return l1_table
        t = tables[0]

        # Adoption gate: only take L2 when it genuinely structures the block
        # better than L1 did.  We measure by distinct, named columns so a mock
        # that just emits generic col_N names never displaces a good L1 header.
        def _distinct_named(cols):
            named = [c for c in cols if c and not re.match(r"^(column|col)_\d+$", c)]
            base = [re.sub(r"_\d+$", "", c) for c in named]
            return len(set(base))

        l1_score = _distinct_named(l1_table.get("columns", []))
        l2_score = _distinct_named(t.get("columns", []))
        if l2_score <= l1_score:
            return l1_table  # L2 did not improve structure -> keep deterministic L1

        # preserve L1 identity/provenance, adopt L2 structure
        t.setdefault("notes", [])
        t["notes"] = list(l1_table.get("notes", [])) + ["L2 fallback (mock LLM)"] + t["notes"]
        t["name"] = l1_table["name"]
        t["sheet"] = l1_table["sheet"]
        t["method"] = "L2"
        return t
    except Exception:
        return l1_table


# ---------------------------------------------------------------------------
# Sheet + workbook drivers
# ---------------------------------------------------------------------------
def parse_sheet(ws, used_names):
    """Extract every clean table from a single worksheet."""
    tables = []
    try:
        grid, merged = _load_grid(ws)
        if not grid:
            return tables
        grid = _unmerge(grid, merged)

        extent = _sheet_extent(grid)
        if extent is None:
            return tables
        r0, r1, c0, c1 = extent

        for (br0, br1) in _split_row_blocks(grid, r0, r1, c0, c1):
            try:
                tbl = _extract_block(
                    grid, br0, br1, c0, c1, ws.title, merged, used_names
                )
            except Exception as exc:  # never let one block kill the sheet
                tbl = None
            if tbl is None:
                continue

            # Keep/drop is decided on the DETERMINISTIC L1 confidence, so a
            # low-quality L2 mock can never resurrect junk that L1 rejected.
            if tbl["confidence"] < DROP_BLOCK_BELOW_CONF:
                continue

            # Only for kept-but-uncertain blocks do we consult L2 to refine.
            if tbl["confidence"] < CONFIDENCE_L2_THRESHOLD:
                tbl = _maybe_l2(grid, br0, br1, c0, c1, tbl)

            tables.append(tbl)
    except Exception:
        # A corrupt/odd sheet must never abort the whole workbook.
        return tables
    return tables


def parse_workbook(path: str) -> list:
    """Parse an .xlsx file into a list of clean table dicts (see CONTRACT.md).

    Robust by construction: a bad sheet is skipped, never raised.
    """
    tables = []
    used_names = set()
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception:
        return tables
    for ws in wb.worksheets:
        tables.extend(parse_sheet(ws, used_names))
    return tables


if __name__ == "__main__":
    import json
    import sys

    target = sys.argv[1] if len(sys.argv) > 1 else "samples/hub_kpi.xlsx"
    out = parse_workbook(target)
    print(f"{target}: {len(out)} table(s)")
    for t in out:
        print(f"  [{t['method']} {t['confidence']}] {t['name']} "
              f"({t['sheet']} {t['cell_range']}) cols={t['columns']}")
        for note in t["notes"]:
            print(f"      - {note}")
