import csv
import io
import re
import uuid
from datetime import datetime, timezone, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.models.organization import Organization
from app.models.membership import Membership
from app.schemas.organization_schema import OrganizationCreate, OrganizationSchema, OrganizationAndRoleSchema, OrganizationUpdate
from app.schemas.organization_schema import MembershipCreate, MembershipSchema, MembershipUpdate, MemberUserCreate
from app.schemas.organization_schema import (
    MembershipImportRow,
    MembershipImportSummary,
    MembershipImportReport,
    MEMBERSHIP_NOTE_MAX_LENGTH,
)
from app.schemas.organization_settings_schema import OrganizationSettingsCreate
from app.services.organization_settings_service import OrganizationSettingsService
from app.schemas.user_schema import UserSchema
from uuid import UUID
from app.models.user import User
from typing import List
from sqlalchemy.orm import selectinload
from fastapi import HTTPException
from sqlalchemy import delete, update
from app.services.llm_service import LLMService
from app.services.test_suite_service import TestSuiteService
from app.settings.config import settings
from fastapi import Request
from typing import Optional
from app.settings.logging_config import get_logger
from app.core.telemetry import telemetry

logger = get_logger(__name__)

# How long an invite link stays valid. Resending rotates the token and resets
# this window.
INVITE_EXPIRY_DAYS = 14


class OrganizationService:

    def __init__(self):
        self.llm_service = LLMService()
        self.organization_settings_service = OrganizationSettingsService()
        self.test_suite_service = TestSuiteService()
    async def create_organization(self, db: AsyncSession, organization_data: OrganizationCreate, current_user: User) -> OrganizationSchema:

        total_orgs = await db.execute(select(Organization))
        total_orgs = total_orgs.scalars().all().__len__()
        if total_orgs > 0 and not settings.dash_config.features.allow_multiple_organizations:
            raise HTTPException(status_code=400, detail="You cannot create more than one organization")
        
        organization = Organization(**organization_data.dict())
        db.add(organization)
        await db.commit()
        await db.refresh(organization)

        # Telemetry: organization created
        try:
            await telemetry.capture(
                "organization_created",
                {
                    "organization_id": str(organization.id),
                    "name_length": len((organization.name or "").strip()),
                },
                user_id=current_user.id,
                org_id=organization.id,
            )
        except Exception:
            pass

        await self.organization_settings_service.create_default_settings(db, organization, current_user)
        await self.add_member(db, MembershipCreate(role="admin", user_id=current_user.id, organization_id=organization.id), current_user)
        await self.llm_service.set_default_models_from_config(db, organization, current_user)
        await self.test_suite_service.ensure_default_for_org(db, organization.id, current_user)

        # Create RBAC role_assignment for the admin system role
        await self._assign_system_role(db, organization.id, str(current_user.id), "admin")

        return OrganizationSchema.from_orm(organization)

    async def _assign_system_role(self, db: AsyncSession, org_id: str, user_id: str, role_name: str) -> None:
        """Assign a system role to a user via role_assignments (RBAC path)."""
        from app.models.role import Role
        from app.models.role_assignment import RoleAssignment

        try:
            result = await db.execute(
                select(Role).where(
                    Role.name == role_name,
                    Role.is_system == True,
                    Role.organization_id.is_(None),
                    Role.deleted_at.is_(None),
                )
            )
            system_role = result.scalar_one_or_none()
            if not system_role:
                return
            # Check if assignment already exists
            existing = await db.execute(
                select(RoleAssignment).where(
                    RoleAssignment.organization_id == org_id,
                    RoleAssignment.role_id == system_role.id,
                    RoleAssignment.principal_type == "user",
                    RoleAssignment.principal_id == user_id,
                )
            )
            if existing.scalar_one_or_none():
                return
            assignment = RoleAssignment(
                organization_id=org_id,
                role_id=system_role.id,
                principal_type="user",
                principal_id=user_id,
            )
            db.add(assignment)
            await db.commit()
        except Exception:
            await db.rollback()
            # Don't break org creation if RBAC tables don't exist yet (pre-migration)

    async def _sync_system_role_assignment(self, db: AsyncSession, org_id: str, user_id: str, role_name: str) -> None:
        """Switch a user's system-role assignment to ``role_name`` (RBAC path).

        Removes any existing system-role (admin/member) assignments for this user
        in this org and adds the one for ``role_name``. Does NOT commit — the
        caller commits as part of its own transaction. Used to keep RBAC in sync
        when the legacy ``update_member`` path changes ``Membership.role``.
        """
        from app.models.role import Role
        from app.models.role_assignment import RoleAssignment

        sys_role_ids = (await db.execute(
            select(Role.id).where(
                Role.is_system == True,
                Role.organization_id.is_(None),
                Role.deleted_at.is_(None),
            )
        )).scalars().all()
        if sys_role_ids:
            await db.execute(
                delete(RoleAssignment).where(
                    RoleAssignment.organization_id == org_id,
                    RoleAssignment.principal_type == "user",
                    RoleAssignment.principal_id == user_id,
                    RoleAssignment.role_id.in_(sys_role_ids),
                )
            )
        target = (await db.execute(
            select(Role).where(
                Role.name == role_name,
                Role.is_system == True,
                Role.organization_id.is_(None),
                Role.deleted_at.is_(None),
            )
        )).scalar_one_or_none()
        if target:
            db.add(RoleAssignment(
                organization_id=org_id,
                role_id=target.id,
                principal_type="user",
                principal_id=user_id,
            ))

    async def get_organization(self, db: AsyncSession, organization_id: str, current_user: User) -> OrganizationSchema:
        result = await db.execute(select(Organization).where(Organization.id == organization_id))
        return result.scalar_one_or_none()
    
    async def get_members(self, db: AsyncSession, organization: Organization, current_user: User) -> List[MembershipSchema]:
        from app.models.role_assignment import RoleAssignment
        from app.models.role import Role
        from app.models.group_membership import GroupMembership
        from app.models.group import Group
        from app.schemas.organization_schema import RoleSummarySchema

        result = await db.execute(
            select(Membership)
            # ★`User.oauth_accounts` is eager-loaded because resolve_auth_origin
            # needs it and a plain attribute read would lazy-load mid-request,
            # which raises under asyncpg. It is the only signal that separates an
            # SSO account from a local one — every account carries a
            # hashed_password, so that cannot be the test.
            .options(selectinload(Membership.user).selectinload(User.oauth_accounts))
            .where(Membership.organization_id == organization.id)
        )
        memberships = result.scalars().all()

        from app.core.permission_resolver import resolve_permissions, FULL_ADMIN
        from app.core.auth_origin import resolve_auth_origin

        schemas = []
        for membership in memberships:
            schema = MembershipSchema.from_orm(membership)
            if membership.user and schema.user is not None:
                # Drives the Sign-in column and whether Set password is offered.
                # A pending invite has no user row and so no origin yet.
                schema.user.auth_origin = resolve_auth_origin(
                    membership.user,
                    oauth_accounts=list(membership.user.oauth_accounts or []),
                )
            if membership.user_id:
                # Registered user: direct ('user' principal) + group-inherited.
                schema.roles = await self._resolve_member_roles(
                    db, organization.id,
                    direct_principal_type="user",
                    direct_principal_id=membership.user_id,
                    group_filter=GroupMembership.user_id == membership.user_id,
                )
                # Derive the coarse `role` label from RBAC so it never disagrees
                # with the member's effective permissions (the legacy
                # Membership.role column drifted on role changes).
                resolved = await resolve_permissions(db, str(membership.user_id), str(organization.id))
                schema.role = "admin" if FULL_ADMIN in resolved.org_permissions else "member"
            else:
                # Pending invite: direct ('membership' principal) +
                # group-inherited via pending group memberships.
                schema.roles = await self._resolve_member_roles(
                    db, organization.id,
                    direct_principal_type="membership",
                    direct_principal_id=membership.id,
                    group_filter=GroupMembership.membership_id == membership.id,
                )
            schemas.append(schema)
        return schemas

    async def _resolve_member_roles(
        self, db: AsyncSession, organization_id: str,
        direct_principal_type: str, direct_principal_id: str, group_filter,
    ) -> List["RoleSummarySchema"]:
        """Resolve a member's direct + group-inherited roles.

        Works for both registered users (principal_type='user', groups joined
        by user_id) and pending invites (principal_type='membership', groups
        joined by membership_id).
        """
        from app.models.role_assignment import RoleAssignment
        from app.models.group_membership import GroupMembership
        from app.models.group import Group
        from app.schemas.organization_schema import RoleSummarySchema

        # Direct role assignments
        ra_result = await db.execute(
            select(RoleAssignment)
            .options(selectinload(RoleAssignment.role))
            .where(
                RoleAssignment.organization_id == organization_id,
                RoleAssignment.principal_type == direct_principal_type,
                RoleAssignment.principal_id == direct_principal_id,
                RoleAssignment.deleted_at.is_(None),
            )
        )
        assignments = ra_result.scalars().all()
        roles = [
            RoleSummarySchema(id=a.role.id, name=a.role.name, source="direct")
            for a in assignments if a.role
        ]

        # Group-inherited role assignments
        gm_result = await db.execute(
            select(GroupMembership.group_id, Group.name)
            .join(Group, Group.id == GroupMembership.group_id)
            .where(
                group_filter,
                Group.organization_id == organization_id,
                GroupMembership.deleted_at.is_(None),
                Group.deleted_at.is_(None),
            )
        )
        member_groups = gm_result.all()  # [(group_id, group_name), ...]

        if member_groups:
            group_ids = [g[0] for g in member_groups]
            group_names = {g[0]: g[1] for g in member_groups}
            seen_role_ids = {r.id for r in roles}

            group_ra_result = await db.execute(
                select(RoleAssignment)
                .options(selectinload(RoleAssignment.role))
                .where(
                    RoleAssignment.organization_id == organization_id,
                    RoleAssignment.principal_type == "group",
                    RoleAssignment.principal_id.in_(group_ids),
                    RoleAssignment.deleted_at.is_(None),
                )
            )
            for a in group_ra_result.scalars().all():
                if a.role and a.role.id not in seen_role_ids:
                    group_name = group_names.get(a.principal_id, "unknown")
                    roles.append(RoleSummarySchema(
                        id=a.role.id,
                        name=a.role.name,
                        source=f"group:{group_name}",
                    ))
                    seen_role_ids.add(a.role.id)

        return roles
    
    async def get_member(self, db: AsyncSession, membership_id: str, organization_id: str, current_user: User) -> MembershipSchema:
        result = await db.execute(
            select(Membership)
            .options(selectinload(Membership.user))
            .where(Membership.id == membership_id, Membership.organization_id == organization_id)
        )
        return result.scalar_one_or_none()
    

    async def _count_org_memberships(self, db: AsyncSession, organization_id) -> int:
        """Count all memberships in an org — active members and pending invites alike.

        Thin wrapper over the shared seat helper (single source of truth), kept for
        the CSV-import projection that counts once and tracks its own running total.
        """
        from app.core.seats import count_org_memberships
        return await count_org_memberships(db, organization_id)

    async def _enforce_user_limit(self, db: AsyncSession, organization_id, adding: int = 1) -> None:
        """Raise 402 if adding `adding` member(s) would exceed the license seat cap.

        No-op when unlicensed/unset (max_users == -1 → unlimited).
        """
        from app.core.seats import enforce_seat_limit
        await enforce_seat_limit(db, organization_id, adding)

    async def add_member(self, db: AsyncSession, membership_data: MembershipCreate, current_user: User) -> MembershipSchema:
        #check if email is already a user
        # if it is, add user_id to membership and remove email
        # then, check if user (or email) already maps to a membership in this organization
        # if it does, raise an error
        # Normalize the invite email to lowercase so it reliably matches the email
        # an SSO/OIDC provider returns at login (Entra/Okta casing is out of our control).
        if membership_data.email:
            membership_data.email = membership_data.email.strip().lower()
        membership_exists = await self._is_email_already_in_organization(db, membership_data.email, membership_data.organization_id)
        if membership_exists:
            raise HTTPException(status_code=400, detail="Already a member with this email")

        # Enforce per-organization seat cap from the enterprise license (if any).
        # Checked after the duplicate guard so re-adding an existing email still
        # 400s rather than being masked by a seat-limit 402.
        await self._enforce_user_limit(db, membership_data.organization_id)
        
        user = await db.execute(
            select(User).where(func.lower(User.email) == (membership_data.email or "").strip().lower())
        )
        user = user.scalar_one_or_none()

        # Store the email for invitation before potentially setting it to None
        invitation_email = membership_data.email

        if user:
            membership_data.user_id = user.id
            membership_data.email = None

        membership = Membership(**membership_data.dict())
        # Pending (unregistered) invite → stamp a 14-day expiry on the invite
        # link. invite_token is auto-generated by the model default.
        if membership.user_id is None:
            membership.invite_expires_at = datetime.utcnow() + timedelta(days=INVITE_EXPIRY_DAYS)

        db.add(membership)
        await db.commit()
        await db.refresh(membership)
        
        # Reload the membership with the user relationship
        result = await db.execute(
            select(Membership)
            .options(selectinload(Membership.user))
            .where(Membership.id == membership.id)
        )
        membership_with_user = result.scalar_one()
        # Telemetry: organization member invited/added
        try:
            await telemetry.capture(
                "organization_member_added",
                {
                    "organization_id": str(membership_with_user.organization_id),
                    "membership_id": str(membership_with_user.id),
                    "role": membership_with_user.role,
                    "user_id": str(membership_with_user.user_id) if membership_with_user.user_id else None,
                },
                user_id=current_user.id,
                org_id=membership_with_user.organization_id,
            )
        except Exception:
            pass
        
        # Send invitation email immediately, but reliably: awaited (so we know
        # the real outcome), retried on transient SMTP errors, and timeout-bounded
        # so a hung relay can't block the request. The outcome is surfaced on the
        # response so the admin UI can warn instead of silently "succeeding".
        # Only pending invites get a sign-up email (existing users already have
        # an account). The link carries the invite token so the recipient can
        # be verified at registration time.
        invite_email_status: Optional[str] = None
        if invitation_email and membership_with_user.user_id is None:
            if not (hasattr(settings, 'email_client') and settings.email_client):
                invite_email_status = "skipped_no_smtp"
            else:
                invite_email_status = await self._send_invitation_email(
                    invitation_email, membership_with_user.invite_token
                )

        # Create RBAC role_assignment if user_id is set
        if membership_with_user.user_id and membership_data.role:
            await self._assign_system_role(db, membership_data.organization_id, membership_with_user.user_id, membership_data.role)

        schema = MembershipSchema.from_orm(membership_with_user)
        schema.invite_email_status = invite_email_status
        return schema

    async def create_member_user(self, db: AsyncSession, data: MemberUserCreate, current_user: User) -> MembershipSchema:
        """Superadmin-only: fully provision a user account (email + password +
        role) and add it to the organization in one step. Unlike add_member
        (invite/self-signup), this creates the User row directly so the account
        is immediately usable — no invite token, no sign-up email.

        The User is created directly (not via UserManager.create) to avoid the
        on_after_register side effect that auto-creates a personal organization.
        """
        email = (data.email or "").strip().lower()

        # Instance-wide uniqueness: one users row per email (accounts merge on
        # lower(email)), so reject if the address already exists anywhere.
        existing = await db.execute(
            select(User).where(func.lower(User.email) == email)
        )
        if existing.scalar_one_or_none() is not None:
            raise HTTPException(status_code=400, detail="A user with this email already exists.")

        # Guard against a duplicate membership (e.g. a pending invite row) in this org.
        membership_exists = await self._is_email_already_in_organization(db, email, data.organization_id)
        if membership_exists:
            raise HTTPException(status_code=400, detail="Already a member with this email")

        # Enforce the per-organization seat cap from the enterprise license (if any).
        await self._enforce_user_limit(db, data.organization_id)

        from fastapi_users.password import PasswordHelper
        ph = PasswordHelper()
        hashed = ph.hash(data.password)

        user = User(
            email=email,
            name=data.name,
            hashed_password=hashed,
            is_active=True,
            is_verified=True,
            is_superuser=False,
        )
        db.add(user)
        await db.commit()
        await db.refresh(user)

        membership = Membership(
            user_id=user.id,
            organization_id=data.organization_id,
            role=data.role,
        )
        db.add(membership)
        await db.commit()
        await db.refresh(membership)

        # Reload the membership with the user relationship
        result = await db.execute(
            select(Membership)
            .options(selectinload(Membership.user))
            .where(Membership.id == membership.id)
        )
        membership_with_user = result.scalar_one()

        # Create RBAC role_assignment if user_id is set
        if membership_with_user.user_id and data.role:
            await self._assign_system_role(db, data.organization_id, membership_with_user.user_id, data.role)

        return MembershipSchema.from_orm(membership_with_user)

    async def get_user_organizations(self, db: AsyncSession, current_user: User) -> List[OrganizationAndRoleSchema]:
        from app.core.permission_resolver import resolve_permissions_bulk

        result = await db.execute(
            select(Organization, Membership.role)
            .join(Membership)
            .where(Membership.user_id == current_user.id)
        )
        results = result.all()
        org_ids = [org.id for org, _ in results]
        # Load settings for these orgs to extract icon_url
        from app.models.organization_settings import OrganizationSettings
        settings_map = {}
        if org_ids:
            sres = await db.execute(select(OrganizationSettings).where(OrganizationSettings.organization_id.in_(org_ids)))
            for s in sres.scalars().all():
                settings_map[s.organization_id] = s

        # Resolve RBAC permissions for ALL orgs in a constant number of queries
        # (was ~3 queries × N orgs, serialized — the dominant whoami cost).
        uid = str(current_user.id)
        resolved_by_org = await resolve_permissions_bulk(db, uid, [str(o) for o in org_ids])

        # License + usage-limit checks are org-independent here — hoist out of the
        # loop so we don't re-check (and, when quotas are enabled, batch) per org.
        from app.ee.license import has_feature
        is_enterprise = has_feature("custom_roles")

        from app.services.usage_policy_service import usage_policy_service
        formatted = []
        for org, role in results:
            icon_url = None
            ai_analyst_name = "AI Analyst"  # Default value
            org_settings = settings_map.get(org.id)
            if org_settings and isinstance(org_settings.config, dict):
                general = org_settings.config.get('general') or {}
                icon_url = general.get('icon_url')
                ai_analyst_name = general.get('ai_analyst_name') or "AI Analyst"

            resolved = resolved_by_org.get(str(org.id))
            if resolved is None:
                from app.core.permission_resolver import ResolvedPermissions
                resolved = ResolvedPermissions()

            # Build resource_permissions dict for frontend
            resource_perms = {}
            for (res_type, res_id), perms in resolved.resource_permissions.items():
                key = f"{res_type}:{res_id}"
                resource_perms[key] = sorted(perms)

            # Derive the coarse `role` label from RBAC (the source of truth) so it
            # can never disagree with the resolved permissions. `role` (from the
            # legacy Membership.role column) was returned verbatim before, which
            # drifted from RBAC on role changes.
            from app.core.permission_resolver import FULL_ADMIN
            derived_role = "admin" if FULL_ADMIN in resolved.org_permissions else "member"

            # Usage quota: when usage_limits is off (the common case) this issues
            # no queries and returns the disabled summary cheaply.
            usage_quota = await usage_policy_service.get_user_quota_summary(
                db, str(org.id), uid,
            )

            formatted.append(OrganizationAndRoleSchema(
                id=org.id,
                name=org.name,
                description=org.description,
                role=derived_role,  # backward compat, now derived from RBAC
                roles=resolved.role_names,
                permissions=sorted(resolved.org_permissions),
                resource_permissions=resource_perms,
                is_enterprise=is_enterprise,
                icon_url=icon_url,
                ai_analyst_name=ai_analyst_name,
                usage_quota=usage_quota,
            ))
        return formatted


    async def remove_member(self, db: AsyncSession, organization_id, membership_id: str, current_user: User, organization: Organization) -> None:
        from app.core.permission_resolver import assert_full_admin_exists

        membership = await self.get_member(db, membership_id, organization_id, current_user)
        if not membership:
            raise HTTPException(status_code=404, detail="Membership not found")
        if membership.user_id:
            # RBAC lockout prevention: ensure at least one user keeps full_admin_access
            await assert_full_admin_exists(db, organization_id, exclude_user_id=membership.user_id)
            # Revoke the standing artifacts that would otherwise let the removed
            # user keep acting in the org without a membership: verified external
            # platform links (Teams/Slack/WhatsApp/email) and their scheduled
            # prompts. The execution-time membership checks are the real gate;
            # this proactively tears down convenience access at removal time and
            # stops departed members' schedules from firing at all.
            await self._revoke_departed_member_access(
                db, organization_id, str(membership.user_id)
            )
        else:
            # Pending invite: clean up any pre-assigned RBAC keyed by this
            # membership. Done explicitly (not via FK cascade) so it holds on
            # SQLite, where foreign-key enforcement is off by default.
            from app.models.role_assignment import RoleAssignment
            from app.models.group_membership import GroupMembership
            from app.models.usage_policy import UsagePolicyAssignment
            await db.execute(
                delete(RoleAssignment).where(
                    RoleAssignment.organization_id == organization_id,
                    RoleAssignment.principal_type == "membership",
                    RoleAssignment.principal_id == membership_id,
                )
            )
            await db.execute(
                delete(GroupMembership).where(GroupMembership.membership_id == membership_id)
            )
            await db.execute(
                delete(UsagePolicyAssignment).where(
                    UsagePolicyAssignment.organization_id == organization_id,
                    UsagePolicyAssignment.principal_type == "membership",
                    UsagePolicyAssignment.principal_id == membership_id,
                )
            )

        await db.execute(delete(Membership).where(Membership.id == membership_id))
        await db.commit()

    async def _revoke_departed_member_access(
        self, db: AsyncSession, organization_id: str, user_id: str
    ) -> None:
        """Tear down a removed user's standing access artifacts in this org.

        - Unverify their external-platform mappings (Teams/Slack/WhatsApp/email)
          so an already-linked chat identity can no longer send queries; the
          next message re-enters the verify flow, which will fail without a
          membership.
        - Deactivate their scheduled prompts on reports in this org and drop the
          APScheduler jobs, so nothing keeps firing as the departed user.

        Best-effort and self-contained: a failure here must not block the
        membership removal itself (the execution-time checks still gate access).
        """
        from app.models.external_user_mapping import ExternalUserMapping
        from app.models.scheduled_prompt import ScheduledPrompt
        from app.models.report import Report

        try:
            await db.execute(
                update(ExternalUserMapping)
                .where(
                    ExternalUserMapping.organization_id == organization_id,
                    ExternalUserMapping.app_user_id == user_id,
                )
                .values(is_verified=False)
            )
        except Exception:
            logger.warning("Failed to unverify external mappings for removed member", exc_info=True)

        sp_ids: list[str] = []
        try:
            result = await db.execute(
                select(ScheduledPrompt.id)
                .join(Report, ScheduledPrompt.report_id == Report.id)
                .where(
                    Report.organization_id == organization_id,
                    ScheduledPrompt.user_id == user_id,
                    ScheduledPrompt.deleted_at == None,
                    ScheduledPrompt.is_active == True,
                )
            )
            sp_ids = [str(r) for r in result.scalars().all()]
            if sp_ids:
                await db.execute(
                    update(ScheduledPrompt)
                    .where(ScheduledPrompt.id.in_(sp_ids))
                    .values(is_active=False)
                )
        except Exception:
            logger.warning("Failed to deactivate scheduled prompts for removed member", exc_info=True)

        await db.commit()

        # Drop the cron jobs after the DB state is committed. Import locally to
        # avoid a module-level dependency on the scheduler from org management.
        for sp_id in sp_ids:
            try:
                from app.services.scheduled_prompt_service import scheduled_prompt_service
                scheduled_prompt_service._remove_job(sp_id)
            except Exception:
                logger.warning(f"Failed to remove cron job for scheduled prompt {sp_id}", exc_info=True)

    async def update_member(self, db: AsyncSession, membership_id: str, organization_id: str, membership_data: MembershipUpdate, current_user: User, organization: Organization) -> MembershipSchema:
        from app.core.permission_resolver import assert_full_admin_exists

        membership = await self.get_member(db, membership_id, organization_id, current_user)
        if not membership:
            raise HTTPException(status_code=404, detail="Membership not found")

        # RBAC lockout prevention: whenever we mutate a membership's role field,
        # verify at least one user still holds full_admin_access in the org.
        if membership.user_id and membership_data.role and membership_data.role != membership.role:
            await assert_full_admin_exists(db, organization_id, exclude_user_id=membership.user_id)

        update = membership_data.dict(exclude_unset=True)
        role_changed = (
            "role" in update
            and update["role"] is not None
            and update["role"] != membership.role
            and membership.user_id is not None
        )
        new_role = update.get("role")
        if "role" in update and update["role"] is not None:
            membership.role = update["role"]
        if "note" in update:
            membership.note = update["note"]
        # Keep RBAC (the source of truth) in sync with the legacy role change so
        # this path can't leave the two stores diverged (a "demoted" admin must
        # actually lose full_admin_access, not just have the string flipped).
        if role_changed:
            await self._sync_system_role_assignment(
                db, organization_id, str(membership.user_id), new_role
            )
        await db.commit()

        # Reload with the user relationship so MembershipSchema serialization
        # doesn't trigger an async lazy-load in a session that's already done.
        result = await db.execute(
            select(Membership)
            .options(selectinload(Membership.user))
            .where(Membership.id == membership.id)
        )
        return MembershipSchema.from_orm(result.scalar_one())

    async def resend_invite(self, db: AsyncSession, membership_id: str, organization_id: str) -> MembershipSchema:
        """Rotate the invite token, reset the 14-day expiry, and re-send the email.

        Pending invites only (a registered member has no invite to resend). The
        old link stops working as soon as the token is rotated.
        """
        membership = await self.get_member(db, membership_id, organization_id, None)
        if not membership:
            raise HTTPException(status_code=404, detail="Membership not found")
        if membership.user_id is not None or not membership.email:
            raise HTTPException(status_code=400, detail="This member has already registered; nothing to resend")

        membership.invite_token = str(uuid.uuid4())
        membership.invite_expires_at = datetime.utcnow() + timedelta(days=INVITE_EXPIRY_DAYS)
        await db.commit()
        await db.refresh(membership)

        status = None
        if hasattr(settings, 'email_client') and settings.email_client:
            status = await self._send_invitation_email(membership.email, membership.invite_token)
        else:
            status = "skipped_no_smtp"

        result = await db.execute(
            select(Membership).options(selectinload(Membership.user)).where(Membership.id == membership.id)
        )
        schema = MembershipSchema.from_orm(result.scalar_one())
        schema.invite_email_status = status
        return schema

    async def get_invite_link(self, db: AsyncSession, membership_id: str, organization_id: str) -> dict:
        """Return the tokenized sign-up link for a pending invite (admin use).

        Lets an admin copy/share the link directly (handy when SMTP is off) and
        is the proof-of-invite the recipient presents at registration. If the
        invite has already expired, the token is regenerated and the 14-day
        window reset so the copied link is always usable (no email is sent —
        that's what Resend does). A still-valid link is returned untouched so we
        don't invalidate one that was already emailed.
        """
        from urllib.parse import quote

        membership = await self.get_member(db, membership_id, organization_id, None)
        if not membership:
            raise HTTPException(status_code=404, detail="Membership not found")
        if membership.user_id is not None or not membership.email:
            raise HTTPException(status_code=400, detail="This member has already registered; no invite link")

        expires = membership.invite_expires_at
        regenerated = False
        if not membership.invite_token or (expires is not None and expires < datetime.utcnow()):
            membership.invite_token = str(uuid.uuid4())
            membership.invite_expires_at = datetime.utcnow() + timedelta(days=INVITE_EXPIRY_DAYS)
            await db.commit()
            await db.refresh(membership)
            regenerated = True

        token = membership.invite_token
        url = (
            f"{settings.dash_config.base_url}/users/sign-up"
            f"?token={quote(token or '')}&email={quote(membership.email)}"
        )
        return {
            "token": token,
            "email": membership.email,
            "url": url,
            "invite_expires_at": membership.invite_expires_at,
            "regenerated": regenerated,
        }

    async def update_organization(self, db: AsyncSession, organization: Organization, data: OrganizationUpdate, current_user: User) -> OrganizationSchema:
        """Update organization basic fields like name/description."""
        update = data.dict(exclude_unset=True)
        if 'name' in update and update['name']:
            organization.name = update['name']
        if 'description' in update:
            organization.description = update['description']
        await db.commit()
        await db.refresh(organization)
        return OrganizationSchema.from_orm(organization)
    
    async def _is_email_already_in_organization(self, db: AsyncSession, email: str, organization_id: str) -> bool:
        email_norm = (email or "").strip().lower()
        user = await db.execute(select(User).where(func.lower(User.email) == email_norm))
        user = user.scalar_one_or_none()
        if user:
            membership = await db.execute(select(Membership).where(Membership.user_id == user.id, Membership.organization_id == organization_id))
            membership = membership.scalar_one_or_none()
            return membership

        email_membership = await db.execute(select(Membership).where(func.lower(Membership.email) == email_norm, Membership.organization_id == organization_id))
        email_membership = email_membership.scalar_one_or_none()
        if email_membership:
            return email_membership 
        
        return False
    
    async def _active_admin_count(self, db: AsyncSession, organization: Organization, current_user: User) -> int:
        """Count active users holding full_admin_access in the org (via RBAC resolver)."""
        from app.core.permission_resolver import resolve_permissions, FULL_ADMIN

        memberships = await self.get_members(db, organization, current_user)
        count = 0
        for m in memberships:
            if not m.user_id:
                continue
            resolved = await resolve_permissions(db, str(m.user_id), str(organization.id))
            if FULL_ADMIN in resolved.org_permissions:
                count += 1
        return count
    
    async def _send_invitation_email(self, email: str, token: Optional[str] = None) -> str:
        """Send the invite email now, reliably. Returns "sent" or "failed".

        Awaited (not fire-and-forget) so the caller knows the real outcome,
        with a couple of retries for transient SMTP blips and a per-attempt
        timeout so a hung relay can't stall the invite request. The link carries
        the invite token (proof of inbox ownership at registration).
        """
        from urllib.parse import quote
        from app.services.notification_service import notification_service
        from app.services.email_copy import invite_email

        params = f"email={quote(email)}"
        if token:
            params = f"token={quote(token)}&{params}"
        sign_up_url = f"{settings.dash_config.base_url}/users/sign-up?{params}"
        subject, body = invite_email(sign_up_url)
        result = await notification_service.send_custom_email(
            recipients=[email],
            subject=subject,
            body=body,
            subtype="plain",
            retries=2,
            timeout=15,
        )
        if result.status != "sent":
            logger.error("Invitation email to %s failed: %s", email, result.error)
        return result.status


    async def get_organization_members(self, db: AsyncSession, current_user: User, organization: Organization) -> List[UserSchema]:
        # should get list of users via membership table
        result = await db.execute(select(Membership).where(Membership.organization_id == organization.id))
        memberships = result.scalars().all()
        user_ids = [membership.user_id for membership in memberships if membership.user_id is not None]
        result = await db.execute(select(User).where(User.id.in_(user_ids)))
        users = result.scalars().all()
        return [UserSchema.from_orm(user) for user in users]

    async def get_organization_groups(self, db: AsyncSession, organization: Organization) -> List[dict]:
        """Minimal group directory for share pickers: id, name, member count."""
        from app.models.group import Group
        from app.models.group_membership import GroupMembership

        groups = (await db.execute(
            select(Group).where(
                Group.organization_id == str(organization.id),
                Group.deleted_at.is_(None),
            ).order_by(Group.name)
        )).scalars().all()
        if not groups:
            return []

        count_rows = (await db.execute(
            select(GroupMembership.group_id, func.count(GroupMembership.id))
            .where(
                GroupMembership.group_id.in_([str(g.id) for g in groups]),
                GroupMembership.deleted_at.is_(None),
            )
            .group_by(GroupMembership.group_id)
        )).all()
        counts = {str(gid): cnt for gid, cnt in count_rows}

        return [
            {
                "id": str(g.id),
                "name": g.name,
                "description": g.description,
                "member_count": counts.get(str(g.id), 0),
            }
            for g in groups
        ]

    # ------------------------------------------------------------------
    # Excel / CSV import of memberships
    # ------------------------------------------------------------------

    MEMBERSHIP_IMPORT_MAX_ROWS = 1000
    _EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

    async def import_members(
        self,
        db: AsyncSession,
        organization: Organization,
        file_bytes: bytes,
        filename: str,
        dry_run: bool,
        current_user: User,
    ) -> MembershipImportReport:
        """Import memberships from .xlsx or .csv.

        Columns: `email` (required), `note` (optional). Other columns are ignored.
        Re-import is additive: existing roles/groups/status are never touched;
        only the `note` field is overwritten. Invites are not re-sent for rows
        whose email already has a (pending or active) membership.
        """
        rows = self._parse_membership_import(file_bytes, filename)
        report_rows: List[MembershipImportRow] = []
        summary = MembershipImportSummary()

        # Enterprise license seat cap. We track a running projection of how many *new*
        # members the import would create so the limit is reported up front — in the
        # dry-run preview and the real run alike — instead of failing row-by-row only
        # once writes start. -1 means unlimited. Existing members (updated/unchanged)
        # don't consume a new seat, so only freshly-created rows count toward it.
        from app.ee.license import get_max_users
        max_users = get_max_users()
        current_members = await self._count_org_memberships(db, organization.id) if max_users >= 0 else 0
        projected_new = 0

        for idx, raw in enumerate(rows, start=2):  # data starts at row 2 (header is row 1)
            email = (raw.get("email") or "").strip()
            note = raw.get("note")
            if note is not None:
                note = str(note).strip() or None
                if note and len(note) > MEMBERSHIP_NOTE_MAX_LENGTH:
                    report_rows.append(MembershipImportRow(
                        row=idx, email=email, note=note, status="error",
                        error=f"Note exceeds {MEMBERSHIP_NOTE_MAX_LENGTH} characters",
                    ))
                    summary.errors += 1
                    continue

            if not email:
                report_rows.append(MembershipImportRow(
                    row=idx, email=None, status="error", error="Missing email",
                ))
                summary.errors += 1
                continue
            if not self._EMAIL_RE.match(email):
                report_rows.append(MembershipImportRow(
                    row=idx, email=email, status="error", error="Invalid email format",
                ))
                summary.errors += 1
                continue

            existing = await self._find_membership_by_email(db, email, organization.id)
            if existing:
                if (existing.note or None) == (note or None):
                    report_rows.append(MembershipImportRow(
                        row=idx, email=email, note=note, status="unchanged",
                    ))
                    summary.unchanged += 1
                else:
                    if not dry_run:
                        existing.note = note
                        await db.flush()
                    report_rows.append(MembershipImportRow(
                        row=idx, email=email, note=note, status="updated",
                    ))
                    summary.updated += 1
                continue

            # New email — invite as a new pending membership.
            # Gate against the license seat cap using the running projection so the
            # overflow is reported identically in dry-run and real runs.
            if max_users >= 0 and current_members + projected_new >= max_users:
                report_rows.append(MembershipImportRow(
                    row=idx, email=email, note=note, status="error",
                    error=(
                        f"User limit reached for your license ({max_users}). "
                        "Contact sales to increase your seat count."
                    ),
                ))
                summary.errors += 1
                continue
            projected_new += 1

            if dry_run:
                report_rows.append(MembershipImportRow(
                    row=idx, email=email, note=note, status="created",
                ))
                summary.created += 1
            else:
                try:
                    await self.add_member(
                        db,
                        MembershipCreate(
                            organization_id=organization.id,
                            email=email,
                            role="member",
                            note=note,
                        ),
                        current_user,
                    )
                    report_rows.append(MembershipImportRow(
                        row=idx, email=email, note=note, status="created",
                    ))
                    summary.created += 1
                except HTTPException as e:
                    report_rows.append(MembershipImportRow(
                        row=idx, email=email, note=note, status="error",
                        error=str(e.detail),
                    ))
                    summary.errors += 1

        if not dry_run:
            await db.commit()

        return MembershipImportReport(dry_run=dry_run, summary=summary, rows=report_rows)

    async def _find_membership_by_email(self, db: AsyncSession, email: str, organization_id: str) -> Optional[Membership]:
        """Find a membership in this org by email, whether the user has registered or not."""
        user_q = await db.execute(select(User).where(User.email == email))
        user = user_q.scalar_one_or_none()
        if user:
            m_q = await db.execute(
                select(Membership).where(
                    Membership.user_id == user.id,
                    Membership.organization_id == organization_id,
                )
            )
            membership = m_q.scalar_one_or_none()
            if membership:
                return membership

        m_q = await db.execute(
            select(Membership).where(
                Membership.email == email,
                Membership.organization_id == organization_id,
            )
        )
        return m_q.scalar_one_or_none()

    def _parse_membership_import(self, file_bytes: bytes, filename: str) -> List[dict]:
        """Parse .xlsx or .csv into a list of {email, note} dicts.

        Header row is required. Column matching is case-insensitive and
        whitespace-trimmed. Columns other than ``email`` and ``note`` are
        ignored. Raises HTTPException(400) on unparseable input or empty file.
        """
        name = (filename or "").lower()
        ext = name.rsplit(".", 1)[-1] if "." in name else ""

        if ext in ("xlsx", "xlsm"):
            try:
                from openpyxl import load_workbook
            except ImportError as e:
                raise HTTPException(status_code=500, detail=f"Excel parsing unavailable: {e}")
            try:
                wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Failed to read xlsx: {e}")
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                raise HTTPException(status_code=400, detail="Empty file")
            header_map = self._build_header_map(header)
            self._require_email_column(header_map)
            data_rows = []
            for raw in rows_iter:
                if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in raw):
                    continue
                data_rows.append({key: raw[idx] if idx < len(raw) else None for key, idx in header_map.items()})
                if len(data_rows) >= self.MEMBERSHIP_IMPORT_MAX_ROWS:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Import exceeds max {self.MEMBERSHIP_IMPORT_MAX_ROWS} rows",
                    )
            return data_rows

        if ext == "csv" or not ext:
            try:
                text = file_bytes.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = file_bytes.decode("latin-1")
            reader = csv.reader(io.StringIO(text))
            try:
                header = next(reader)
            except StopIteration:
                raise HTTPException(status_code=400, detail="Empty file")
            header_map = self._build_header_map(header)
            self._require_email_column(header_map)
            data_rows = []
            for raw in reader:
                if not any((c or "").strip() for c in raw):
                    continue
                data_rows.append({key: raw[idx] if idx < len(raw) else None for key, idx in header_map.items()})
                if len(data_rows) >= self.MEMBERSHIP_IMPORT_MAX_ROWS:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Import exceeds max {self.MEMBERSHIP_IMPORT_MAX_ROWS} rows",
                    )
            return data_rows

        raise HTTPException(status_code=400, detail=f"Unsupported file type: .{ext}")

    @staticmethod
    def _build_header_map(header_row) -> dict:
        """Map known column names (case-insensitive) to their column indexes."""
        mapping = {}
        for idx, name in enumerate(header_row or []):
            if name is None:
                continue
            key = str(name).strip().lower()
            if key in ("email", "note") and key not in mapping:
                mapping[key] = idx
        return mapping

    @staticmethod
    def _require_email_column(header_map: dict) -> None:
        if "email" not in header_map:
            raise HTTPException(
                status_code=400,
                detail="Missing required 'email' column in header row",
            )
