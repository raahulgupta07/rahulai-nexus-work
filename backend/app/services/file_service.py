import asyncio
import os

from fastapi import UploadFile
from sqlalchemy.orm import Session
from app.schemas.file_schema import FileSchema, FileSchemaWithMetadata, FileSchemaWithCompletionId
from app.models.file import File
import uuid
from app.models.report import Report
from app.models.user import User
from app.models.organization import Organization
from app.models.sheet_schema import SheetSchema
from typing import Optional
from datetime import datetime
from fastapi import HTTPException
from app.models.file import report_file_association
from app.models.data_source_file_association import data_source_file_association
from app.models.data_source import DataSource
from app.models.file_tag import FileTag
from sqlalchemy.ext.asyncio import AsyncSession
import aiofiles
from sqlalchemy import select, exists
from app.core.telemetry import telemetry
from app.services.file_preview import generate_file_preview
import logging

logger = logging.getLogger(__name__)


# ── intake decision record ──────────────────────────────────────────────────
# The smart-intake librarian decides where every uploaded file belongs, with a
# confidence and a one-line reason. Both were logged and thrown away, so a badge
# in the UI could not be checked: a well-founded verdict and a coin-flip looked
# the same. These helpers keep that record on the File row.
#
# Stored under File.preview["intake"] rather than in a new column — preview is
# already a nullable JSON column, so this needs no migration. That does mean the
# preview writer and this writer share one column, hence merge_preview below;
# see the note at its call site in upload_file.
INTAKE_PREVIEW_KEY = "intake"

# The paths this file contributed to `connection.config.file_paths`, so removing
# the file can remove exactly those and nothing else.
#
# A plain CSV is reflected under its own managed path, so it needs no record.
# An .xlsx does: it is expanded into one CSV per sheet, each named with a FRESH
# uuid4 (`excel_ingest.xlsx_to_csvs` line 22), and the raw .xlsx is deliberately
# never added to file_paths. Nothing on the File row points at those sheet CSVs,
# so without this record a deleted spreadsheet leaves every one of its tables
# alive and unreachable — the same defect as the CSV case, minus any way to find
# the wreckage. Files uploaded before this shipped have no record; delete falls
# back to matching the file's own path and says what it could not remove.
DERIVED_PATHS_PREVIEW_KEY = "derived_paths"


def merge_intake_into_preview(preview, intake):
    """Return ``preview`` with the intake record preserved.

    ``generate_file_preview`` builds a fresh dict from the file's bytes and
    knows nothing about the intake decision, so assigning its result directly
    would drop a record written earlier in the same request. Callers that
    replace ``File.preview`` wholesale must route through here.
    """
    base = dict(preview) if isinstance(preview, dict) else {}
    if intake:
        base[INTAKE_PREVIEW_KEY] = intake
    return base


def read_intake_decision(file_row):
    """The stored verdict for a file, or None when it predates this record."""
    preview = getattr(file_row, "preview", None)
    if not isinstance(preview, dict):
        return None
    intake = preview.get(INTAKE_PREVIEW_KEY)
    return intake if isinstance(intake, dict) else None


def merge_derived_paths_into_preview(preview, paths):
    """Return ``preview`` carrying the paths this file contributed."""
    base = dict(preview) if isinstance(preview, dict) else {}
    if paths:
        base[DERIVED_PATHS_PREVIEW_KEY] = list(paths)
    return base


def carry_forward_preview_records(fresh_preview, previous_preview):
    """Rebuild a regenerated preview without losing the records stored beside it.

    `preview` holds two things that are NOT derived from the file's bytes — the
    intake verdict and the derived-path list — so anything that regenerates the
    preview must carry them across explicitly. Listed by key rather than merged
    wholesale so a stale preview cannot resurrect fields the new one dropped.
    """
    out = dict(fresh_preview) if isinstance(fresh_preview, dict) else {}
    if isinstance(previous_preview, dict):
        for key in (INTAKE_PREVIEW_KEY, DERIVED_PATHS_PREVIEW_KEY):
            if key in previous_preview:
                out[key] = previous_preview[key]
    return out


def owned_table_paths(file_row):
    """Every ``file_paths`` entry that exists because of this file.

    The file's own managed path is included whether or not it was recorded — a
    CSV is always reflected under it, and including it costs nothing when it was
    never there. Recorded derived paths are added for spreadsheets.
    """
    paths = []
    own = getattr(file_row, "path", None)
    if own:
        import os as _os

        paths.append(_os.path.abspath(own))
    preview = getattr(file_row, "preview", None)
    if isinstance(preview, dict):
        for p in preview.get(DERIVED_PATHS_PREVIEW_KEY) or []:
            if isinstance(p, str) and p and p not in paths:
                paths.append(p)
    return paths


async def _record_intake_decision(
    db: AsyncSession,
    file_id: Optional[str],
    *,
    destination: Optional[str],
    confidence: float,
    reason: str,
    decided_by: str,
) -> None:
    """Store how this file's destination was chosen.

    Best-effort by design: this is a record ABOUT the routing, never a step in
    it. A failure here must not cost the user their upload, so every path
    swallows and logs.
    """
    if not file_id:
        return
    try:
        row = (await db.execute(select(File).filter(File.id == file_id))).scalar_one_or_none()
        if row is None:
            return
        try:
            conf = max(0.0, min(1.0, float(confidence)))
        except (TypeError, ValueError):
            conf = 0.0
        record = {
            "destination": destination,
            "confidence": round(conf, 2),
            "reason": (reason or "").strip(),
            # "llm" — the librarian read the content. "deterministic" — decided
            # by shape/extension, either as the cheap fast path or because the
            # LLM call failed. The UI should not present the two identically.
            "decided_by": decided_by,
            "decided_at": datetime.utcnow().isoformat(),
        }
        # Reassign rather than mutate: SQLAlchemy does not track in-place edits
        # to a JSON column, and the write would be silently dropped.
        row.preview = merge_intake_into_preview(row.preview, record)
        db.add(row)
        await db.commit()
    except Exception as err:
        logger.warning(f"could not record intake decision for file {file_id}: {err}")


def scope_files_to_user_uploads(all_files, data_sources, enabled: bool = True):
    """Focus scoping for the agent's readable file space.

    When the user has uploaded their OWN files to a report (not files
    auto-snapshotted onto the report from a bound agent's data source), the
    agent should focus on those uploads instead of also reading every bound
    agent's inherited knowledge files. In Auto mode a report binds every agent,
    so without this a "summarize this attached file" turn drags in all agents'
    files (Abbott.pptx, Definitions.xlsx, CRM Q&A.docx, …).

    A "user upload" is any file in ``all_files`` whose id is NOT present in any
    bound data source's ``files`` snapshot. When at least one such upload exists
    AND ``enabled`` is True, returns just the uploads; otherwise returns
    ``all_files`` unchanged. Single source of truth — reused by agent_v2,
    FilesContextBuilder, and the read_file / grep_files session-file resolvers.

    Never raises: any failure returns ``all_files`` unchanged (fail-open).
    """
    try:
        all_files = list(all_files or [])
        if not enabled or not all_files:
            return all_files
        snapshot_ids = set()
        for _ds in (data_sources or []):
            for _f in (getattr(_ds, "files", None) or []):
                _fid = getattr(_f, "id", None)
                if _fid is not None:
                    snapshot_ids.add(str(_fid))
        user_uploaded = [f for f in all_files if str(getattr(f, "id", "")) not in snapshot_ids]
        if user_uploaded:
            return user_uploaded
        return all_files
    except Exception as _scope_err:
        logger.warning(
            "scope_files_to_user_uploads: focus filter failed, using all files: %s",
            _scope_err,
        )
        return list(all_files or [])


class FileService:
    def __init__(self):
        pass

    async def upload_file(
        self,
        db: AsyncSession,
        file: UploadFile,
        current_user: User,
        organization: Organization,
        report_id: Optional[str] = None,
        data_source_id: Optional[str] = None,
        learn: bool = True,
    ) -> FileSchema:
        # Generate a unique filename to prevent overwriting existing files
        unique_filename = f"{uuid.uuid4()}_{file.filename}"
        file_location = f"uploads/files/{unique_filename}"

        # The image pre-creates uploads/files, but a volume mounted over
        # uploads/ hides it: Docker seeds a fresh named volume from the image,
        # Kubernetes mounts an empty PVC/emptyDir and shadows the subdirectory.
        # Without this every upload raised FileNotFoundError under k8s.
        os.makedirs(os.path.dirname(file_location), exist_ok=True)

        # Async file writing
        async with aiofiles.open(file_location, "wb") as buffer:
            content = await file.read()
            await buffer.write(content)

        # Create the database entry
        db_file = File(
            filename=file.filename,
            content_type=file.content_type,
            path=file_location,
            user_id=current_user.id,
            organization_id=organization.id
        )

        db.add(db_file)
        await db.commit()
        await db.refresh(db_file)

        # Telemetry: file uploaded (minimal fields only)
        try:
            await telemetry.capture(
                "file_uploaded",
                {
                    "file_id": str(db_file.id),
                    "content_type": db_file.content_type,
                    "bytes": len(content or b""),
                    "report_id": report_id,
                    "data_source_id": data_source_id,
                },
                user_id=current_user.id,
                org_id=organization.id,
            )
        except Exception:
            pass

        # Associate with report if provided
        if report_id:
            stmt = select(Report).filter(Report.id == report_id)
            result = await db.execute(stmt)
            report = result.scalar_one_or_none()

            if report:
                report.files.append(db_file)
                await db.commit()
                await db.refresh(report)

        # Associate with data source if provided
        if data_source_id:
            stmt = select(DataSource).filter(
                DataSource.id == data_source_id,
                DataSource.organization_id == organization.id,
            )
            result = await db.execute(stmt)
            data_source = result.scalar_one_or_none()

            if not data_source:
                raise HTTPException(status_code=404, detail="Data source not found")

            data_source.files.append(db_file)
            await db.commit()
            await db.refresh(data_source)

            # If this is a CSV data source and the uploaded file is a CSV,
            # reflect the (server-generated) managed storage path into the
            # connection's config.file_paths so the CSV client can build a
            # DuckDB table from it, then trigger a schema refresh. This is
            # entirely best-effort and defensive: any failure here must NOT
            # break the upload — the file stays attached regardless.
            try:
                import os
                import json

                # Resolve the data source's single connection (M:N list).
                conn = None
                connections = getattr(data_source, "connections", None) or []
                if connections:
                    conn = connections[0]

                # Determine whether the uploaded file looks like a CSV.
                _fname = (file.filename or "").lower()
                _ctype = (file.content_type or "").lower()
                _csv_content_types = {
                    "text/csv",
                    "application/csv",
                    "application/vnd.ms-excel",  # some browsers send this for .csv
                    "text/plain",
                }
                _looks_csv = _fname.endswith(".csv") or _ctype in _csv_content_types

                # Detect xlsx: each sheet is expanded into its own CSV, since
                # DuckDB read_csv_auto cannot read the raw .xlsx binary.
                _xlsx_content_types = {
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-excel",
                }
                _looks_xlsx = _fname.endswith(".xlsx") or _ctype in _xlsx_content_types

                # Server-generated absolute path — never derived from user input.
                abs_path = os.path.abspath(file_location)

                # Build the list of CSV paths to reflect into the connection.
                # Plain CSV: the file itself. xlsx: one CSV per non-empty sheet
                # (do NOT append the raw .xlsx — DuckDB can't read it).
                paths_to_add = []
                if _looks_xlsx:
                    out_dir = os.path.dirname(abs_path)
                    try:
                        from app.services.excel_convert import convert_xlsx
                        paths_to_add = convert_xlsx(abs_path, out_dir)
                    except Exception:
                        paths_to_add = []
                elif _looks_csv:
                    paths_to_add = [abs_path]

                _is_def = False
                _smart_handled = False

                # ── Smart file intake (flag-gated) ──────────────────────────
                # An LLM "librarian" READS the file content and decides where it
                # goes (table | instruction | skill | knowledge) — decision comes
                # from content, not the filename. instruction/skill are rewritten
                # into clean structured rows; knowledge docs are chunked into the
                # retrievable metadata index; table falls through to the existing
                # DuckDB path. Any failure falls back to the legacy path.
                #
                # Flag OFF => this method is never called, so behavior is
                # byte-identical to the pre-smart-intake path.
                try:
                    from app.settings.config import settings as _isettings
                    if getattr(_isettings, "smart_file_intake", False):
                        _res = await self._smart_file_intake(
                            db,
                            abs_path=abs_path,
                            fname=_fname,
                            ctype=_ctype,
                            looks_xlsx=_looks_xlsx,
                            looks_csv=_looks_csv,
                            data_source=data_source,
                            current_user=current_user,
                            organization=organization,
                            file_id=str(db_file.id),
                        )
                        if _res.get("handled"):
                            _smart_handled = True
                            _is_def = True  # handled → do not also build a table
                except Exception as _si_err:
                    logger.warning(f"smart_file_intake failed for '{_fname}', falling back to legacy: {_si_err}")
                    _smart_handled = False

                # Legacy: Definitions/dictionary file → ONE instruction (raw dump),
                # skip table-building. Only runs when smart intake did NOT handle it.
                if not _smart_handled:
                    try:
                        from app.services.def_ingest import is_definitions_file, xlsx_to_definitions_block
                        if is_definitions_file(_fname) and _looks_xlsx:
                            block = xlsx_to_definitions_block(abs_path)
                            if block:
                                from app.services.instruction_service import InstructionService
                                from app.schemas.instruction_schema import InstructionCreate
                                instr = InstructionCreate(
                                    text=block,
                                    category="data_modeling",
                                    kind="instruction",
                                    status="published",
                                    load_mode="always",
                                    data_source_ids=[str(data_source.id)],
                                )
                                try:
                                    await InstructionService().create_instruction(
                                        db, instr, current_user, organization, auto_finalize=True,
                                    )
                                    _is_def = True  # handled as instruction; do not also build a table
                                except Exception:
                                    _is_def = False
                    except Exception:
                        _is_def = False

                # If it became an instruction, don't also reflect it into file_paths as a table.
                if _is_def:
                    paths_to_add = []

                if conn is not None and getattr(conn, "type", None) == "csv" and paths_to_add:
                    # Connection config is stored as a JSON string (matches
                    # create_data_source Mode-1: config=json.dumps(config)).
                    cfg = conn.config
                    if isinstance(cfg, str):
                        cfg = json.loads(cfg) if cfg else {}
                    cfg = cfg or {}

                    existing = [
                        p.strip()
                        for p in (cfg.get("file_paths") or "").splitlines()
                        if p.strip()
                    ]
                    for _p in paths_to_add:
                        if _p not in existing:
                            existing.append(_p)
                    cfg["file_paths"] = "\n".join(existing)
                    conn.config = json.dumps(cfg)
                    db.add(conn)
                    # Record what this file put into file_paths so removing the
                    # file can take exactly those entries back out. Matters most
                    # for .xlsx, whose per-sheet CSVs carry fresh uuids that
                    # nothing else on this row could ever match.
                    try:
                        db_file.preview = merge_derived_paths_into_preview(
                            db_file.preview, paths_to_add
                        )
                    except Exception:
                        pass
                    # This file's data now lives as a queryable DuckDB table, so
                    # mark it table-backing. The agent should query the table, not
                    # also read the raw CSV — reading both risks double-counting or
                    # picking a stale copy. Table-backing files are excluded from
                    # the read_file / knowledge catalog (see File.is_agent_readable).
                    try:
                        db_file.source_kind = "table_backing"
                        db.add(db_file)
                    except Exception:
                        pass
                    await db.commit()

                    # Make the new file queryable, FAST. Previously this awaited
                    # llm_sync in-request (schema refresh + 3 LLM generations =
                    # ~32s hang on the upload). Split it:
                    #   a) refresh_data_source_schema synchronously (no LLM) so
                    #      the reflected table is visible immediately — the create
                    #      wizard's Review step still sees tables right away.
                    #   b) regenerate the onboarding overview in the BACKGROUND
                    #      (schedule_overview_relearn — fire-and-forget, own
                    #      session, swallows errors), so the response returns fast.
                    # Both best-effort; a failure here must never break the upload.
                    try:
                        from app.services.data_source_service import DataSourceService
                        _dss = DataSourceService()
                        await _dss.refresh_data_source_schema(
                            db, str(data_source.id), organization, current_user
                        )
                        # Background learn only when the agent opts into LLM
                        # learning AND the caller asked to learn (learn=False is
                        # used to batch: learn once after the LAST file of a
                        # multi-file upload). The schema refresh above always runs.
                        if learn and getattr(data_source, "use_llm_sync", True):
                            _dss.schedule_overview_relearn(
                                str(data_source.id),
                                str(current_user.id) if current_user else None,
                                str(organization.id),
                            )
                    except Exception as _sync_err:
                        logger.warning(
                            f"CSV schema sync after upload failed for data source "
                            f"{data_source.id}: {_sync_err}"
                        )
            except Exception as _reflect_err:
                # Reflection is non-fatal — the file remains attached.
                logger.warning(
                    f"Failed to reflect uploaded CSV into connection file_paths: {_reflect_err}"
                )

        # Generate raw preview (no LLM) - fast, instant
        try:
            # Earlier in this same request, intake recorded its verdict and the
            # reflect step recorded which file_paths entries this file created —
            # both under keys on `preview`. generate_file_preview builds a fresh
            # dict from the file's bytes and knows nothing about either, so
            # assigning it directly would erase both on every upload that
            # produces a preview, silently.
            db_file.preview = carry_forward_preview_records(
                generate_file_preview(db_file), db_file.preview
            )
            db.add(db_file)
            await db.commit()
            await db.refresh(db_file)
            logger.info(f"Generated preview for file {db_file.filename} (type: {db_file.preview.get('type', 'unknown') if db_file.preview else 'none'})")
        except Exception as e:
            # Preview generation failure is non-fatal - log and continue
            logger.warning(f"Failed to generate preview for {db_file.filename}: {e}")
        
        # Return the file schema
        file_schema = FileSchema.from_orm(db_file)

        return file_schema

    async def _smart_file_intake(
        self,
        db: AsyncSession,
        *,
        abs_path: str,
        fname: str,
        ctype: str,
        looks_xlsx: bool,
        looks_csv: bool,
        data_source: DataSource,
        current_user: User,
        organization: Organization,
        file_id: Optional[str] = None,
        force_destination: Optional[str] = None,
        keep_existing: bool = False,
    ) -> dict:
        """LLM-driven "librarian": read the file content and route it.

        ``force_destination`` overrides the verdict with the caller's choice and
        skips the classifier entirely — no LLM call, no confidence to report.
        The rewriters below still run, so a forced instruction is written up
        properly rather than pasted in raw. Every existing caller omits it and
        behaves exactly as before.

        A forced conversion also RETIRES whatever this file produced last time,
        unless ``keep_existing`` says otherwise — converting is a correction, and
        leaving the superseded artifacts behind means the agent reads both the
        old filing and the new one.

        Reads the uploaded file's extracted content and asks an LLM to pick the
        destination store (table | instruction | skill | knowledge). The verdict
        comes from CONTENT, not the filename. The deterministic ``classify_file``
        is used only as a cheap fast-path for an obviously-clean tabular grid and
        as a fallback when the LLM call fails.

        Routing:
          * table       -> return ``{"handled": False}`` so the caller builds the
                           DuckDB table (unchanged behavior).
          * instruction -> rewrite into data-dictionary / rule instructions.
          * skill       -> draft a ``kind="skill"`` instruction.
          * knowledge   -> chunk into retrievable ``metadata_resources`` rows.

        Returns ``{"handled": bool, "destination": str|None, "created": int}``.
        Never raises for a routing/LLM failure — falls back so the upload always
        succeeds.
        """
        import asyncio as _aio
        from app.services.file_classifier import classify_file, classify_file_llm
        from app.services import file_normalizer as _norm
        from app.services.instruction_service import InstructionService as _IS
        from app.schemas.instruction_schema import InstructionCreate as _IC

        # Deterministic pass — cheap fast-path + fallback + shape signal.
        _cls = classify_file(abs_path, fname, ctype) or {}
        _det_dest = _cls.get("destination")
        try:
            _det_conf = float(_cls.get("confidence") or 0.0)
        except (TypeError, ValueError):
            _det_conf = 0.0

        # How the verdict below was reached, recorded onto the File row at the
        # end of this block. Held in locals because the LLM verdict is produced
        # inside a try/except and the fast-path skips it entirely.
        _decided_by = "deterministic"
        _decided_conf = _det_conf
        _decided_reason = str(_cls.get("reason") or "")

        # Gather the file's content ONCE (sheet preview OR document text).
        _raw = ""
        if looks_xlsx or looks_csv:
            try:
                from app.services.def_ingest import xlsx_to_definitions_block as _x2d
                _raw = _x2d(abs_path) or ""
            except Exception:
                _raw = ""
        else:
            try:
                from app.data_sources.clients._document_text import extract_document_text as _edt
                _raw = _edt(abs_path, name=fname) or ""
            except Exception:
                _raw = ""

        # Build an LLM infer() callable (offloaded to a worker thread —
        # LLM.inference's usage check can't run under an active event loop).
        from app.ai.llm import LLM as _LLM
        from app.dependencies import async_session_maker as _asm
        _model = await organization.get_default_llm_model(db)
        _llm = _LLM(_model, usage_session_maker=_asm)

        def _infer(_p):
            return _llm.inference(_p, usage_scope="file_intake")

        # ---- decide the destination (LLM verdict preferred) ----
        _dest = None
        # An explicit choice ends the decision. Recorded as decided_by="user" so
        # the UI never presents a person's instruction as a machine's guess, and
        # so a later automatic pass can tell which files it must not re-file.
        if force_destination in ("table", "instruction", "skill", "knowledge"):
            _dest = force_destination
            _decided_by = "user"
            _decided_conf = 1.0
            _decided_reason = "Chosen by the user."
        # Fast-path: an obviously-clean rectangular grid stays a table without an
        # LLM round-trip.
        elif _det_dest == "table" and _det_conf >= 0.85:
            _dest = "table"
        else:
            if _raw and _raw.strip():
                try:
                    _v = await _aio.to_thread(classify_file_llm, _raw, fname, ctype, _infer)
                    if _v and _v.get("destination") in ("table", "instruction", "skill", "knowledge"):
                        _dest = _v["destination"]
                        _decided_by = "llm"
                        try:
                            _decided_conf = float(_v.get("confidence") or 0.0)
                        except (TypeError, ValueError):
                            _decided_conf = 0.0
                        _decided_reason = str(_v.get("reason") or "")
                        logger.info(
                            f"smart_file_intake: LLM librarian '{fname}' -> {_dest} "
                            f"(conf={_v.get('confidence')}; {_v.get('reason','')})"
                        )
                except Exception as _lerr:
                    logger.warning(f"smart_file_intake: LLM classifier failed for '{fname}': {_lerr}")
            if _dest is None:
                _dest = _det_dest  # deterministic fallback
                logger.info(f"smart_file_intake: fallback deterministic '{fname}' -> {_dest}")

        # Persist the verdict BEFORE acting on it, so the record exists even for
        # the destinations that return early below. Until now this decision was
        # written to the container log and discarded, which left a correct call
        # and a wrong one looking identical in the UI.
        await _record_intake_decision(
            db, file_id,
            destination=_dest,
            confidence=_decided_conf,
            reason=_decided_reason,
            decided_by=_decided_by,
        )

        # Table (or nothing usable) → let the existing table/legacy path handle it.
        if _dest in (None, "table"):
            return {"handled": False, "destination": _dest, "created": 0}

        # Back-link so the files API can derive each file's "fate": stamp the
        # source file id onto every artifact this file produces (ai_source on
        # instructions/skills; raw_data.source_file_id on knowledge chunks).
        _ai_src = (f"file:{file_id}" if file_id else None)

        # A conversion REPLACES what this file previously produced. Without
        # this, converting Knowledge to Instruction leaves all the knowledge
        # chunks in place and adds rules beside them, and pressing convert twice
        # stacks a second copy of everything — proven live on a probe file that
        # ended up carrying both an instruction and a skill from one document.
        # Only forced conversions withdraw; an ordinary re-classification is not
        # a decision to discard anything.
        if force_destination and file_id and not keep_existing:
            await self._withdraw_file_artifacts(db, file_id)

        _created = 0
        if _dest == "skill" and _raw:
            _spec = await _aio.to_thread(_norm.draft_skill_from_doc, _raw, _infer)
            _steps = "\n".join(f"- {s}" for s in (_spec.get("steps") or [])) or (_spec.get("body") or "")
            _body = f"SKILL: {_spec.get('name','skill')}\nWhen to use: {_spec.get('when_to_use','')}\nSteps:\n{_steps}"
            await _IS().create_instruction(
                db, _IC(text=_body, category="system", kind="skill", status="published", load_mode="always", ai_source=_ai_src, data_source_ids=[str(data_source.id)]),
                current_user, organization, auto_finalize=True,
            )
            _created += 1
        elif _dest == "instruction" and _raw:
            _shape = (_cls.get("signals") or {}).get("l2_shape")
            if _shape == "term_meaning" or looks_xlsx or looks_csv:
                _text = await _aio.to_thread(_norm.normalize_definitions, _raw, _infer)
                await _IS().create_instruction(
                    db, _IC(text=_text or _raw, category="data_modeling", kind="instruction", status="published", load_mode="always", ai_source=_ai_src, data_source_ids=[str(data_source.id)]),
                    current_user, organization, auto_finalize=True,
                )
                _created += 1
            else:
                # Rule/logic doc → consolidate into ONE 'intelligent' instruction
                # (was up to 20 separate 'always' rows → context bloat). One
                # bulleted instruction, capped, pulled only when relevant.
                _rules = [r.strip() for r in (await _aio.to_thread(_norm.extract_rules_from_doc, _raw, _infer) or []) if r and r.strip()]
                _rules = _rules[:12]
                if _rules:
                    _stem = (fname.rsplit(".", 1)[0] or "rules").strip()
                    _body = f"Rules from {_stem}:\n" + "\n".join(f"- {_r}" for _r in _rules)
                    await _IS().create_instruction(
                        db, _IC(text=_body, category="general", kind="instruction", status="published", load_mode="intelligent", ai_source=_ai_src, data_source_ids=[str(data_source.id)]),
                        current_user, organization, auto_finalize=True,
                    )
                    _created += 1
        elif _dest == "knowledge":
            from app.services.knowledge_ingest import ingest_knowledge_document
            _created += await ingest_knowledge_document(
                db,
                abs_path=abs_path,
                filename=fname,
                content_type=ctype,
                data_source=data_source,
                organization=organization,
                source_text=_raw or None,
                source_file_id=file_id,
            )

        if _created:
            # Hide the raw file from the agent now that its content lives as
            # instructions/skills or knowledge chunks — the agent should use the
            # derived artifacts, not re-read the source doc (mirrors the
            # table_backing dedup). This method only runs when smart_file_intake
            # is ON, and File.is_agent_readable only honors these kinds under the
            # same flag, so a flag-off instance never sees these values.
            try:
                _new_kind = "knowledge_backing" if _dest == "knowledge" else "instruction_backing"
                if file_id:
                    _frow = (await db.execute(select(File).filter(File.id == file_id))).scalar_one_or_none()
                    if _frow is not None:
                        _frow.source_kind = _new_kind
                        db.add(_frow)
                        await db.commit()
            except Exception as _sk_err:
                logger.warning(f"smart_file_intake: could not stamp source_kind for '{fname}': {_sk_err}")
            logger.info(f"smart_file_intake: '{fname}' -> {_dest} ({_created} item(s)) for ds {data_source.id}")
            return {"handled": True, "destination": _dest, "created": _created}
        return {"handled": False, "destination": _dest, "created": 0}

    async def reingest_file(
        self,
        db: AsyncSession,
        file_id: str,
        data_source_id: str,
        organization: Organization,
        current_user: User,
        destination: Optional[str] = None,
        keep_existing: bool = False,
    ) -> dict:
        """Re-run LLM classification + routing on an already-uploaded file.

        With ``destination`` this becomes a conversion: the file is routed where
        the caller says, skipping the classifier. Without it the classifier runs
        again and may reach the same verdict as last time — which is why the UI
        offers "convert to X" rather than a bare "try again".

        Shares the exact intake logic used at upload time (``_smart_file_intake``)
        so a file that was uploaded before this feature existed — or that should
        be re-sorted — can be (re)routed to table / instruction / skill /
        knowledge without re-uploading. Flag-gated: a no-op (``skipped``) when
        ``smart_file_intake`` is OFF.
        """
        import os as _os

        # Flag gate — behavior byte-identical (nothing happens) when OFF.
        from app.settings.config import settings as _isettings
        if not getattr(_isettings, "smart_file_intake", False):
            return {"handled": False, "skipped": "smart_file_intake disabled"}

        if destination is not None and destination not in (
            "table", "instruction", "skill", "knowledge"
        ):
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Unknown destination '{destination}'. "
                    "Expected one of: table, instruction, skill, knowledge."
                ),
            )

        # Resolve + org-scope the data source.
        ds_stmt = select(DataSource).filter(
            DataSource.id == data_source_id,
            DataSource.organization_id == organization.id,
        )
        data_source = (await db.execute(ds_stmt)).scalar_one_or_none()
        if not data_source:
            raise HTTPException(status_code=404, detail="Data source not found")

        # File must belong to this org AND be associated with the data source.
        file_stmt = select(File).filter(
            File.id == file_id,
            File.organization_id == organization.id,
        )
        db_file = (await db.execute(file_stmt)).scalar_one_or_none()
        if not db_file:
            raise HTTPException(status_code=404, detail="File not found")

        assoc_stmt = select(data_source_file_association).filter_by(
            data_source_id=data_source_id, file_id=file_id
        )
        if not (await db.execute(assoc_stmt)).first():
            raise HTTPException(status_code=404, detail="File is not associated with this data source")

        # Only user uploads can be re-ingested (a table-backing copy is already a
        # queryable table — re-routing it would double-handle its data).
        if (getattr(db_file, "source_kind", "upload") or "upload") == "table_backing":
            return {"handled": False, "skipped": "file is table-backing", "file_id": file_id}

        abs_path = _os.path.abspath(db_file.path)
        if not _os.path.exists(abs_path):
            raise HTTPException(status_code=404, detail="File content not found on disk")

        fname = (db_file.filename or "").lower()
        ctype = (db_file.content_type or "").lower()
        _csv_content_types = {"text/csv", "application/csv", "application/vnd.ms-excel", "text/plain"}
        _xlsx_content_types = {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        }
        looks_csv = fname.endswith(".csv") or ctype in _csv_content_types
        looks_xlsx = fname.endswith(".xlsx") or ctype in _xlsx_content_types

        try:
            result = await self._smart_file_intake(
                db,
                abs_path=abs_path,
                fname=fname,
                ctype=ctype,
                looks_xlsx=looks_xlsx,
                looks_csv=looks_csv,
                data_source=data_source,
                current_user=current_user,
                organization=organization,
                file_id=file_id,
                force_destination=destination,
                keep_existing=keep_existing,
            )
        except HTTPException:
            # A deliberate 4xx from inside the intake (an unusable destination
            # for this file type, say) must reach the user as itself, not be
            # relabelled a 500 by the catch-all below.
            raise
        except Exception as _err:
            logger.warning(f"reingest_file failed for '{fname}' ({file_id}): {_err}")
            raise HTTPException(status_code=500, detail=f"Re-ingest failed: {_err}")

        result["file_id"] = file_id
        return result

    async def save_bytes_as_file(
        self,
        db: AsyncSession,
        content: bytes,
        filename: str,
        content_type: str,
        current_user: User,
        organization: Organization,
        report_id: Optional[str] = None,
        completion_id: Optional[str] = None,
    ) -> File:
        """Persist raw bytes (e.g. an inbound email attachment) as a report File.

        Mirrors ``upload_file`` but takes bytes instead of an ``UploadFile``:
        writes to disk, creates the row, optionally links to a report, and
        generates a preview. Returns the ``File`` ORM object.
        """
        safe_name = os.path.basename(filename or "attachment") or "attachment"
        unique_filename = f"{uuid.uuid4()}_{safe_name}"
        file_location = f"uploads/files/{unique_filename}"

        # Same shadowed-mount hazard as upload_file — see the note there.
        os.makedirs(os.path.dirname(file_location), exist_ok=True)

        async with aiofiles.open(file_location, "wb") as buffer:
            await buffer.write(content)

        db_file = File(
            filename=safe_name,
            content_type=content_type or "application/octet-stream",
            path=file_location,
            user_id=current_user.id,
            organization_id=organization.id,
        )
        db.add(db_file)
        await db.commit()
        await db.refresh(db_file)

        if report_id:
            if completion_id:
                # Associate with the report but tag with the completion, so the
                # file is available to the agent (report.files -> <files> context
                # and read_file) while staying hidden from the user's composer
                # attachment tray (which hides completion-tagged files).
                await db.execute(
                    report_file_association.insert().values(
                        report_id=report_id, file_id=db_file.id, completion_id=completion_id
                    )
                )
                await db.commit()
                await db.refresh(db_file)
            else:
                stmt = select(Report).filter(Report.id == report_id)
                result = await db.execute(stmt)
                report = result.scalar_one_or_none()
                if report:
                    report.files.append(db_file)
                    await db.commit()
                    await db.refresh(report)

        try:
            db_file.preview = generate_file_preview(db_file)
            db.add(db_file)
            await db.commit()
            await db.refresh(db_file)
        except Exception as e:  # noqa: BLE001 — preview is best-effort
            logger.warning(f"Failed to generate preview for {db_file.filename}: {e}")

        return db_file
    
    async def remove_file_from_report(self, db: AsyncSession, file_id: str, report_id: str, organization: Organization, current_user: User):
        stmt = select(Report).filter(Report.id == report_id)
        result = await db.execute(stmt)
        report = result.scalar_one_or_none()
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")

        stmt = select(File).filter(File.id == file_id)
        result = await db.execute(stmt)
        file = result.scalar_one_or_none()
        if not file:
            raise HTTPException(status_code=404, detail="File not found")

        stmt = select(report_file_association).filter_by(
            report_id=report_id, file_id=file_id
        )
        result = await db.execute(stmt)
        association = result.first()
        if not association:
            raise HTTPException(status_code=404, detail="File is not associated with this report")

        await db.execute(
            report_file_association.delete().where(
                (report_file_association.c.report_id == report_id) &
                (report_file_association.c.file_id == file_id)
            )
        )
        await db.commit()

        return True
        
    async def get_files(self, db: AsyncSession, organization: Organization):
        stmt = select(File).filter(File.organization_id == organization.id)
        result = await db.execute(stmt)
        files = result.scalars().all()

        # get files with tags
        for file in files:
            stmt = select(FileTag).filter(FileTag.file_id == file.id)
            result = await db.execute(stmt)
            file.tags = result.scalars().all()

            stmt = select(SheetSchema).filter(SheetSchema.file_id == file.id)
            result = await db.execute(stmt)
            file.schemas = result.scalars().all()

        return files

    async def get_files_by_report(self, db: AsyncSession, report_id: str, organization: Organization):
        from app.models.report_data_source_association import report_data_source_association

        stmt = select(Report).filter(Report.id == report_id)
        result = await db.execute(stmt)
        report = result.scalar_one_or_none()
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")

        # Query files with completion_id from the association table
        stmt = (
            select(File, report_file_association.c.completion_id)
            .join(report_file_association, File.id == report_file_association.c.file_id)
            .where(report_file_association.c.report_id == report_id)
        )
        result = await db.execute(stmt)
        rows = result.all()

        # File ids inherited from any of the report's data sources. We
        # treat any overlap as inherited — collisions with user-uploaded
        # files are vanishingly unlikely in practice.
        inherited_stmt = (
            select(data_source_file_association.c.file_id)
            .join(
                report_data_source_association,
                data_source_file_association.c.data_source_id ==
                report_data_source_association.c.data_source_id,
            )
            .where(report_data_source_association.c.report_id == report_id)
        )
        inherited_res = await db.execute(inherited_stmt)
        inherited_ids = {str(r[0]) for r in inherited_res.all()}

        # Build response with completion_id and inheritance flag included
        files_with_completion = []
        for file, completion_id in rows:
            file_dict = FileSchema.from_orm(file).dict()
            file_dict['completion_id'] = str(completion_id) if completion_id else None
            file_dict['from_data_source'] = str(file.id) in inherited_ids
            files_with_completion.append(FileSchemaWithCompletionId(**file_dict))

        return files_with_completion

    async def get_files_by_data_source(
        self,
        db: AsyncSession,
        data_source_id: str,
        organization: Organization,
    ) -> list[FileSchema]:
        ds_stmt = select(DataSource).filter(
            DataSource.id == data_source_id,
            DataSource.organization_id == organization.id,
        )
        ds_result = await db.execute(ds_stmt)
        data_source = ds_result.scalar_one_or_none()
        if not data_source:
            raise HTTPException(status_code=404, detail="Data source not found")

        files_stmt = (
            select(File)
            .join(data_source_file_association, File.id == data_source_file_association.c.file_id)
            .where(data_source_file_association.c.data_source_id == data_source_id)
        )
        files_res = await db.execute(files_stmt)
        files = files_res.scalars().all()

        # Derive each file's "fate" (what the smart intake did with it) so the UI
        # can badge it. table_backing wins outright; otherwise we look for an
        # artifact this file produced via the back-links stamped at intake time:
        #   * an instruction/skill row with ai_source == "file:{id}"
        #   * a knowledge metadata_resource with raw_data.source_file_id == id
        # No match => "upload" (parked, not ingested). Best-effort: any failure
        # leaves fate derived from source_kind alone.
        instr_file_ids: set[str] = set()
        knowledge_file_ids: set[str] = set()
        file_ids = [str(f.id) for f in files]
        try:
            if file_ids:
                from app.models.instruction import Instruction
                from app.models.metadata_resource import MetadataResource

                _tags = [f"file:{fid}" for fid in file_ids]
                # Only LIVE (non-soft-deleted) instructions count as this file's
                # fate — a file whose instructions were all deleted but which has
                # live knowledge chunks must derive fate="knowledge", not
                # "instruction" from a stale leftover.
                _ai_rows = (await db.execute(
                    select(Instruction.ai_source).where(
                        Instruction.organization_id == organization.id,
                        Instruction.ai_source.in_(_tags),
                        Instruction.deleted_at.is_(None),
                    )
                )).all()
                for (_src,) in _ai_rows:
                    if _src and _src.startswith("file:"):
                        instr_file_ids.add(_src[len("file:"):])

                _mr_rows = (await db.execute(
                    select(MetadataResource.raw_data).where(
                        MetadataResource.data_source_id == data_source_id,
                        MetadataResource.resource_type == "knowledge",
                        MetadataResource.deleted_at.is_(None),
                    )
                )).all()
                for (_rd,) in _mr_rows:
                    if isinstance(_rd, dict):
                        _sfid = _rd.get("source_file_id")
                        if _sfid:
                            knowledge_file_ids.add(str(_sfid))
        except Exception as _fate_err:
            logger.warning(f"fate derivation failed for ds {data_source_id}: {_fate_err}")

        out: list[FileSchema] = []
        for f in files:
            schema = FileSchema.from_orm(f)
            fid = str(f.id)
            _sk = (getattr(f, "source_kind", "upload") or "upload")
            # Fast path: the source_kind stamped at intake is authoritative.
            # instruction_backing / knowledge_backing map straight to their fate;
            # the ai_source / raw_data back-link queries below stay as a fallback
            # for legacy rows ingested before source_kind was stamped.
            if _sk == "table_backing":
                schema.fate = "table_backing"
            elif _sk == "instruction_backing":
                schema.fate = "instruction"
            elif _sk == "knowledge_backing":
                schema.fate = "knowledge"
            elif fid in instr_file_ids:
                schema.fate = "instruction"
            elif fid in knowledge_file_ids:
                schema.fate = "knowledge"
            else:
                schema.fate = "upload"
            # How that fate was arrived at. None for files ingested before the
            # decision was recorded — the UI shows the badge alone for those
            # rather than inventing a reason for a call it cannot account for.
            schema.intake = read_intake_decision(f)
            out.append(schema)
        return out

    async def remove_file_from_data_source(
        self,
        db: AsyncSession,
        file_id: str,
        data_source_id: str,
        organization: Organization,
        current_user: User,
    ):
        # `connections` is eager-loaded because the withdrawal below reads it.
        # A DataSource has no `.type` of its own — the connector type lives on
        # the connection — and touching that relationship lazily inside an async
        # session raises MissingGreenlet rather than loading it.
        from sqlalchemy.orm import selectinload

        ds_stmt = (
            select(DataSource)
            .filter(
                DataSource.id == data_source_id,
                DataSource.organization_id == organization.id,
            )
            .options(selectinload(DataSource.connections))
        )
        ds_result = await db.execute(ds_stmt)
        data_source = ds_result.scalar_one_or_none()
        if not data_source:
            raise HTTPException(status_code=404, detail="Data source not found")

        assoc_stmt = select(data_source_file_association).filter_by(
            data_source_id=data_source_id, file_id=file_id
        )
        assoc_result = await db.execute(assoc_stmt)
        if not assoc_result.first():
            raise HTTPException(status_code=404, detail="File is not associated with this data source")

        db_file = (await db.execute(select(File).filter(File.id == file_id))).scalar_one_or_none()

        await db.execute(
            data_source_file_association.delete().where(
                (data_source_file_association.c.data_source_id == data_source_id) &
                (data_source_file_association.c.file_id == file_id)
            )
        )
        await db.commit()

        # Upload writes FOUR things; for a long time this undid one of them.
        #
        # Reflecting a CSV's path into `connection.config.file_paths` is what
        # makes it a queryable table (see upload_file). Deleting only the
        # association left that path in place, so the table stayed active and
        # the agent kept answering from a file the user had removed — with
        # nothing anywhere reporting a problem. Everything below is the inverse
        # of that reflect, and every step is best-effort: a file the user asked
        # to remove must come off the list even if the cleanup behind it fails.
        removed_paths: list[str] = []
        if db_file is not None:
            try:
                removed_paths = await self._withdraw_file_tables(
                    db, db_file, data_source, organization, current_user
                )
            except Exception as err:
                logger.warning(
                    f"could not withdraw tables for removed file {file_id}: {err}"
                )

            # The File row itself outlived every delete before this, accumulating
            # rows that belong to no data source and appear in no list.
            try:
                db_file.deleted_at = datetime.utcnow()
                db.add(db_file)
                await db.commit()
            except Exception as err:
                logger.warning(f"could not soft-delete file row {file_id}: {err}")

        return {"success": True, "removed_paths": removed_paths}

    async def _retire_all_connection_tables(self, db: AsyncSession, conn, data_source) -> int:
        """Stand down every table this file connection produced.

        Only called when the connection has no files left, where "no tables" is
        a fact we just wrote rather than an introspection result that might be a
        transient failure. Deactivates the agent's tables and deletes the
        catalog rows behind them, so nothing is left advertising data that no
        longer exists.
        """
        from sqlalchemy import delete as _sql_delete, update as _sql_update
        from app.models.connection_table import ConnectionTable
        from app.models.datasource_table import DataSourceTable

        retired = 0
        try:
            conn_table_ids = [
                str(row.id) for row in (await db.execute(
                    select(ConnectionTable).filter(
                        ConnectionTable.connection_id == str(conn.id)
                    )
                )).scalars().all()
            ]
            if conn_table_ids:
                # Stand the agent's tables down AND unlink them, in one write.
                #
                # The unlink is not tidiness: `datasource_tables.connection_table_id`
                # is a foreign key with no ON DELETE rule, so deleting the catalog
                # rows underneath a live reference raises
                # `fk_datasource_tables_connection_table_id` and — because this
                # whole method is best-effort — the error would be swallowed and
                # the table left standing, which is the exact bug being fixed.
                doomed = [
                    str(row.id) for row in (await db.execute(
                        select(DataSourceTable).where(
                            DataSourceTable.datasource_id == data_source.id,
                            DataSourceTable.connection_table_id.in_(conn_table_ids),
                        )
                    )).scalars().all()
                ]
                retired = len(doomed)

                if doomed:
                    # The agent's table rows are DELETED, not just deactivated.
                    # Deactivating leaves them listed on the Tables tab — the
                    # paginated reader does not filter soft-deleted rows either,
                    # so a table whose file is gone would keep appearing forever,
                    # greyed out and unexplainable.
                    #
                    # Three of the four things that reference these rows have no
                    # ON DELETE rule (checked against the live schema:
                    # table_stats, table_usage_events, table_feedback_events all
                    # NO ACTION; user_data_source_tables is SET NULL and looks
                    # after itself). They describe a table that no longer exists,
                    # so they go with it — and if they did not, the delete would
                    # raise and this best-effort method would swallow it, leaving
                    # exactly the ghost row it exists to remove.
                    from app.models.table_stats import TableStats
                    from app.models.table_usage_event import TableUsageEvent
                    from app.models.table_feedback_event import TableFeedbackEvent

                    for model, column in (
                        (TableStats, TableStats.datasource_table_id),
                        (TableUsageEvent, TableUsageEvent.datasource_table_id),
                        (TableFeedbackEvent, TableFeedbackEvent.datasource_table_id),
                    ):
                        await db.execute(_sql_delete(model).where(column.in_(doomed)))

                    await db.execute(
                        _sql_delete(DataSourceTable).where(DataSourceTable.id.in_(doomed))
                    )
                # The catalog rows go last: the domain rows above reference them,
                # and that foreign key has no ON DELETE rule either.
                await db.execute(
                    _sql_delete(ConnectionTable).where(
                        ConnectionTable.id.in_(conn_table_ids)
                    )
                )
            await db.commit()
            if retired:
                logger.info(
                    f"retired {retired} table(s) for connection {conn.id} — no files remain"
                )
        except Exception as err:
            logger.warning(f"could not retire tables for connection {conn.id}: {err}")
        return retired

    async def _withdraw_file_artifacts(self, db: AsyncSession, file_id: str) -> dict:
        """Retire the instructions, skills and knowledge chunks a file produced.

        Both stores are reached through the SAME back-links the files API uses to
        derive a file's badge — instructions carry ``ai_source="file:{id}"`` and
        knowledge chunks carry ``raw_data.source_file_id``. Using the same links
        matters: anything this misses keeps its badge, so a file would show as
        both Knowledge and Instruction with no way to tell which the agent is
        actually reading.

        Soft-deleted rather than deleted — a conversion is a judgement call and
        the previous filing may be the better one. Best-effort throughout: a
        failure here must not cost the user the conversion they asked for.
        """
        retired = {"instructions": 0, "knowledge": 0}
        marker = f"file:{file_id}"
        try:
            from app.models.instruction import Instruction

            rows = (await db.execute(
                select(Instruction).filter(
                    Instruction.ai_source == marker,
                    Instruction.deleted_at.is_(None),
                )
            )).scalars().all()
            for row in rows:
                row.deleted_at = datetime.utcnow()
                db.add(row)
            retired["instructions"] = len(rows)
        except Exception as err:
            logger.warning(f"could not retire instructions for file {file_id}: {err}")

        try:
            from app.models.metadata_resource import MetadataResource

            rows = (await db.execute(
                select(MetadataResource).filter(
                    MetadataResource.resource_type == "knowledge",
                    MetadataResource.deleted_at.is_(None),
                )
            )).scalars().all()
            hits = [
                r for r in rows
                if isinstance(getattr(r, "raw_data", None), dict)
                and str(r.raw_data.get("source_file_id") or "") == str(file_id)
            ]
            for row in hits:
                row.deleted_at = datetime.utcnow()
                # Retrieval filters on is_active as well as deleted_at depending
                # on the path, so both are cleared — a chunk that is soft-deleted
                # but still active would keep being served.
                row.is_active = False
                db.add(row)
            retired["knowledge"] = len(hits)
        except Exception as err:
            logger.warning(f"could not retire knowledge chunks for file {file_id}: {err}")

        try:
            await db.commit()
        except Exception as err:
            logger.warning(f"could not commit artifact withdrawal for file {file_id}: {err}")
            return {"instructions": 0, "knowledge": 0}

        if retired["instructions"] or retired["knowledge"]:
            logger.info(
                f"retired {retired['instructions']} instruction(s) and "
                f"{retired['knowledge']} knowledge chunk(s) superseded for file {file_id}"
            )
        return retired

    async def _withdraw_file_tables(
        self,
        db: AsyncSession,
        db_file: File,
        data_source: DataSource,
        organization: Organization,
        current_user: User,
    ) -> list[str]:
        """Take a removed file's paths back out of the connection, and re-sync.

        Returns the paths actually removed, so the caller can report what it did
        rather than claim a cleanup that found nothing to clean.

        The re-sync is what retires the table, and it is deliberately NOT done by
        hand here: `refresh_schema` already deletes ConnectionTable rows that the
        connection no longer reports, and `sync_domain_tables_from_connection`
        already stands down the domain tables pointing at them. Reimplementing
        that would mean maintaining a second, quietly diverging copy of the
        prune. Note the prune only runs when the credential resolve is
        authoritative — true for `csv` connections, which are `system_only`.
        """
        import json as _json
        import os as _os

        owned = owned_table_paths(db_file)
        if not owned:
            return []

        conn = None
        for candidate in (getattr(data_source, "connections", None) or []):
            if getattr(candidate, "type", None) == "csv":
                conn = candidate
                break
        if conn is None:
            return []

        cfg = conn.config
        if isinstance(cfg, str):
            cfg = _json.loads(cfg) if cfg else {}
        cfg = cfg or {}

        existing = [p.strip() for p in (cfg.get("file_paths") or "").splitlines() if p.strip()]
        # Compare on the absolute path: entries were written via os.path.abspath
        # at upload, but a config edited by hand (or by an older build) may not
        # have been, and a near-miss here silently leaves the table alive.
        owned_set = {_os.path.abspath(p) for p in owned}
        kept = [p for p in existing if _os.path.abspath(p) not in owned_set]
        removed = [p for p in existing if _os.path.abspath(p) in owned_set]
        if not removed:
            return []

        cfg["file_paths"] = "\n".join(kept)
        # Reassign rather than mutate — SQLAlchemy does not track in-place edits
        # to a JSON column and the write would be dropped at commit.
        conn.config = _json.dumps(cfg)
        db.add(conn)
        await db.commit()

        from app.services.data_source_service import DataSourceService

        await DataSourceService().refresh_data_source_schema(
            db, str(data_source.id), organization, current_user
        )

        # Removing the LAST file needs its own step, because the refresh above
        # deliberately refuses to prune down to nothing.
        #
        # `refresh_schema` bails at "No tables returned from get_schemas()"
        # BEFORE it reaches the prune (connection_service.py:1332). That guard is
        # right for a database: an introspection that comes back empty usually
        # means the connection broke, and letting that wipe the shared catalog
        # would turn a blip into data loss. But it cannot tell "the source is
        # unreachable" from "the source is genuinely empty", so on a file
        # connection with no files left it leaves the final table behind —
        # active, queryable, and backed by nothing. Observed live: five files
        # removed cleanly, the sixth left `mm_conso_data_report_mar_25` standing.
        #
        # Emptiness is not inferred here. We wrote the empty path list a few
        # lines above, so for THIS connection zero files provably means zero
        # tables. Narrow to that case on purpose — the guard keeps protecting
        # every partial removal, which is already proven to prune correctly.
        if not kept:
            await self._retire_all_connection_tables(db, conn, data_source)

        logger.info(
            f"withdrew {len(removed)} table path(s) for removed file "
            f"'{db_file.filename}' from data source {data_source.id}"
        )
        return removed

    # ==========================================================================
    # DEPRECATED: LLM-based schema extraction methods
    # These methods are no longer called during file upload.
    # We now use raw preview generation instead (see generate_file_preview).
    # Kept for backward compatibility and potential manual re-processing.
    # ==========================================================================
    
    async def _create_sheet_schemas_legacy(self, db: AsyncSession, file: File, model):
        """
        DEPRECATED: LLM-based Excel schema extraction.
        
        This method uses LLM to extract structured schema from Excel files.
        It is no longer called during upload - we now use raw previews instead.
        Kept for backward compatibility if manual schema extraction is needed.
        """
        import warnings
        warnings.warn(
            "_create_sheet_schemas_legacy is deprecated. Use generate_file_preview instead.",
            DeprecationWarning,
            stacklevel=2
        )
        
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
        import xlrd
        from app.ai.agents.excel import ExcelAgent
        
        sheet_names = []
        workbook = None

        if file.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            try:
                workbook = load_workbook(filename=file.path, read_only=True)
                sheet_names = workbook.sheetnames
            except InvalidFileException as e:
                raise HTTPException(status_code=400, detail=f"Failed to process .xlsx file: {e}")
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Error processing Excel file: {e}")
        
        elif file.content_type == "application/vnd.ms-excel":
            try:
                workbook = xlrd.open_workbook(filename=file.path)
                sheet_names = workbook.sheet_names()
            except xlrd.XLRDError as e:
                raise HTTPException(status_code=400, detail=f"Failed to process .xls file: {e}")
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Error processing Excel file: {e}")

        if not sheet_names:
            return 0

        try:
            processed_sheets_count = 0
            for index, sheet_name in enumerate(sheet_names):
                ea = ExcelAgent(file, model)
                # Offload — `get_schema` ultimately calls sync
                # `LLM.inference` which can't run its usage-limit check
                # from an active event loop without a wired `loop`.
                schema = await asyncio.to_thread(ea.get_schema, index)

                if schema and "sheet_name" in schema:
                    sc = SheetSchema(
                        sheet_name=schema["sheet_name"],
                        sheet_index=index,
                        schema=schema,
                        file_id=file.id
                    )
                    db.add(sc)
                    processed_sheets_count += 1

            if processed_sheets_count > 0:
                await db.commit()
            return processed_sheets_count
        
        except Exception as e:
            await db.rollback()
            raise HTTPException(status_code=500, detail=f"Error processing sheet schemas: {e}")

        finally:
            if hasattr(workbook, 'close') and callable(workbook.close):
                workbook.close()

    async def _process_pdf_legacy(self, db: AsyncSession, file: File, model):
        """
        DEPRECATED: LLM-based PDF tag extraction.
        
        This method uses LLM to extract semantic tags from PDF files.
        It is no longer called during upload - we now use raw text preview instead.
        Kept for backward compatibility if manual tag extraction is needed.
        """
        import warnings
        warnings.warn(
            "_process_pdf_legacy is deprecated. Use generate_file_preview instead.",
            DeprecationWarning,
            stacklevel=2
        )
        
        import tiktoken
        from app.ai.agents.doc.doc import DocAgent
        
        da = DocAgent(file, model)
        content = da.get_content()

        tags = []   
        tokenizer = tiktoken.get_encoding("cl100k_base")

        tokens = tokenizer.encode(content)
        chunk_size = 100000
        overlap = 300

        for i in range(0, len(tokens), chunk_size - overlap):
            chunk = tokenizer.decode(tokens[i:i+chunk_size])
            # Offload to a thread — `get_tags_from_text` calls sync
            # `LLM.inference`, whose pre-call usage-limit check raises
            # when invoked from a running event loop with no `loop` set.
            new_tags = await asyncio.to_thread(da.get_tags_from_text, chunk, tags)
            tags.extend(new_tags)
        
        file_tags = []
        
        for tag in tags:
            file_tag = FileTag(
                key=tag["tag"],
                value=tag["value"],
                file_id=file.id
            )
            file_tags.append(file_tag)
        
        for file_tag in file_tags:
            db.add(file_tag)
        await db.commit()
        
        return tags


    async def create_or_get_report_file_association(self, db: AsyncSession, report_id: str, file_id: str):
        # 1. Fetch Report and File
        report_stmt = select(Report).where(Report.id == report_id)
        report_result = await db.execute(report_stmt)
        report = report_result.scalar_one_or_none()
        if not report:
            raise HTTPException(status_code=404, detail=f"Report not found: {report_id}")

        file_stmt = select(File).where(File.id == file_id)
        file_result = await db.execute(file_stmt)
        file = file_result.scalar_one_or_none()
        if not file:
            raise HTTPException(status_code=404, detail=f"File not found: {file_id}")

        # 2. Check if association already exists (more efficient check)
        # Assuming 'files' is the relationship attribute on the Report model
        # Adjust if the relationship is defined differently
        association_exists_stmt = select(exists().where(
            report_file_association.c.report_id == report_id,
            report_file_association.c.file_id == file_id
        ))
        association_exists = await db.scalar(association_exists_stmt)

        # 3. If not associated, create the association by appending
        if not association_exists:
            try:
                # Append the file to the report's collection. SQLAlchemy handles the insert.
                # Ensure the relationship is correctly defined in your models
                # (e.g., on Report: files = relationship("File", secondary=report_file_association, backref="reports"))
                # If the relationship is defined on the File model instead (e.g., file.reports.append(report)), use that.
                report.files.append(file) 
                db.add(report) # Add the modified report to the session if needed
                await db.commit()
                await db.refresh(report) # Refresh report to potentially load the updated relationship
                print(f"Association created between Report {report_id} and File {file_id}")
                return True # Indicate association was created
            except Exception as e:
                await db.rollback()
                print(f"Error creating association: {e}") # Log the specific error
                # Consider raising a specific exception or HTTPException
                raise HTTPException(status_code=500, detail=f"Failed to create association: {e}")
        else:
            print(f"Association already exists between Report {report_id} and File {file_id}")
            return False # Indicate association already existed
        